import pandas as pd
import numpy as np
import os
import csv
import math
import sys
import difflib
import newViewSequence_1_11
from difflib import SequenceMatcher
from Bio import SeqIO
from Bio.Blast import NCBIWWW
from Bio.Blast import NCBIXML
from os import listdir
from collections import OrderedDict
from pandas import ExcelWriter
from pandas import ExcelFile
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
import urllib.request, urllib.parse, urllib.error
from timeit import default_timer as timer
import time
import signal
import datetime
from shutil import copyfile

runfiles = []
df = pd.DataFrame()
dfloc = ''
foldername = ''
try:
    folders = os.listdir('files')
except:
    folders = ""
files = []
output_file = ''
logtxt = ''
filename = ''
filename_ndx = 0
position = 0
HCC = False
adj_name = None
check = False
check_rc = None
check_rcseq = None
check_rcname = None

summary = pd.DataFrame
supporting = []
major = []
summeaningfulreads = 0
sumuniquereads = 0
summajorjunctionreads = 0

sampletype = []
sample = []
reads = []
HBVreads = []
percentHBVreads = []
chimericreads = []
consideredreads = []
meaningfulreads = []
uniquereads = []
supportingreads = []
majorjunctionreads = []

cumulativegenes = []
cumulativejunctions = []

mj = []
driver = []

hlc = []
vlc = []
hrc = []
vrc = []
rn = []
rl = []
sequence = []
chrom = []
gene = []
ig = []
dtg = []
gd = []
fcs = []
go = []
fo = []
hl = []
ho = []
va = []
vl = []
vo = []
htm = []
htma = []
vtm = []
vtma = []
ol = []
oltm = []
oltma = []
insert = []
mh = []
hmf = []
vmf = []
hmq = []
vmq = []
fqs = []
index = []

ho1 = []
vo1 = []
hlc1 = []
vlc1 = []

rc = []
rcseq = []
rclength = []
rc1 = []
rc_name = []

ndr = [] #noduplicatereads
ndrseq = []
ndrlength = []
ndr1 = []

meaningful = [] #meaningfulreads
meaningfulseq = []
meaningfullength = []
meaningfulcount = []
meaningful1 = []

unique = [] #no supporting reads
uniqueseq = []
uniquelength = []
unique1 = []
uniquecount = []
uniqueTUP = []

suppreads = []

searched = []

drivergenes = []

tempseq = ''

complement = {'A': 'T', 'C': 'G', 'G': 'C', 'T': 'A'} 

sname = []
sample_names = []
sample_names_adj = []
sample_names_hcc = []

inputFilesFolder = ""

def signal_handler(signum, frame):
    raise Exception("Timed out!")

def checkRequirements(parent):
    status = True
    if not os.path.isfile("HBV to index ration analysis_finished pooled library.xlsx"):
        parent.logInfo("Error: file 'HBV to index ration analysis_finished pooled library.xlsx' is not found")
        status = False
    if os.path.exists("files"):
        if os.path.isfile("files"):
              parent.logInfo("Error: folder 'files' is not found")
              parent.logInfo("       create 'files' folder and put testfiles into 'files' folder")
              status = False
    else:
        parent.logInfo("Error: folder 'files' is not found")
        parent.logInfo("       create 'files' folder and put testfiles into 'files' folder")
        status = False
    return status
    

def process(parent, filesFolder):

    global sname, sample_names_adj, sample_names_hcc, sample_dict, sample_names, folders, inputFilesFolder 
    
    # old_stdout = sys.stdout
    # cwd = os.getcwd()
    # log_file = open(cwd + '/files/summary_log.txt',"w")
    # sys.stdout = log_file

    start = timer()
    inputFilesFolder = filesFolder
    if (inputFilesFolder == ""):
        cwd = os.getcwd().replace("\\", "/")
        inputFilesFolder = cwd + "/files/"
    else:
        inputFilesFolder += "/"
        try:
            folders = os.listdir(inputFilesFolder)
        except:
            folders = ""
    parent.logInfo("processing files from "+inputFilesFolder+" folder\n")


    sample_info_df = pd.read_excel('HBV to index ration analysis_finished pooled library.xlsx')
    sample_names = sample_info_df['Name'].tolist()[0:140]
    for name in sample_names:
            sname.append(''.join(ch for ch in str(name) if ch.isalnum()))
    sample_types = sample_info_df['sample type'].tolist()[0:140]
    sample_dict = dict(zip(sample_names, sample_types))

    for x in range(0, len(sample_types)):
            if 'adjacent' in sample_types[x] or 'AJC' in sample_types[x]:
                    # matches = difflib.get_close_matches(sname[x], sname, 2, cutoff = 0.2)
                    sample_names_adj.append(sample_names[x]) #adj name
                    # sample_names_hcc.append(sample_names[sname.index(matches[1])]) #hcc match
            elif 'HCC' in sample_types[x] and 'library' not in sample_types[x] and 'urine' not in sample_types[x] and 'before' not in sample_types[x] and 'lung' not in sample_types[x]:
                    sample_names_hcc.append(sample_names[x]) #hcc match

    # get a blank database
    copyfile("DriverGeneDatabase_original.xlsx", "DriverGeneDatabase.xlsx")
    
    def reverse_complement(seq):
            newseq = ''
            for letter in seq:
                    try:
                            newseq += complement[letter]
                    except:
                            newseq += 'N'
            return newseq

    """
    def markdups(f, samname):
            samfile_in = pysam.AlignmentFile(os.getcwd() + '/files/' + foldername + '/' + samname, "r")
            samfile_out = pysam.Samfile(os.getcwd() + '/files/' + foldername + '/[new]' + samname, "w", template=samfile_in)
            print ('reading IJ56.R_1__hostAlign.sam ... \n')
            total = 0
            mapped = 0
            for read in samfile_in.fetch():
                    total += 1
                    if not read.is_unmapped:
                            mapped += 1
                            # print ('query name = ' + read.query_name)
                            rname = read.query_name
                            try:
                                    tempindex = rname.index('.')
                                    read.query_name = rname[tempindex+1:-2]
                            except:
                                    pass
                            samfile_out.write(read)

            samfile_in.close()
            samfile_out.close()

            command1 = 'java -jar /Users/amylu/Desktop/picard-2.10.9/picard.jar SortSam I=' + os.getcwd() + '/files/'+foldername+'/[new]'+samname+' O='+os.getcwd() + '/files/'+foldername+'/[sorted]'+ samname + ' SORT_ORDER=queryname'
            command2 = 'samtools rmdup -S ' + os.getcwd() + '/files/'+foldername+'/[sorted]'+samname+' ' + os.getcwd() + '/files/'+foldername+'/[rmdup]'+ samname

            os.system(command1)
            os.system(command2)

            print ('\ntotal reads : ' + str(total))
            print ('mapped reads : ' + str(mapped))

            unique_read_names = []
            samfile = pysam.AlignmentFile(os.getcwd() + '/files/'+foldername+'/[rmdup]'+samname, "r")
            for read in samfile.fetch():
                    rname = read.query_name
                    unique_read_names.append(rname)

            removed = 0

            df = pd.read_csv(cwd+'/files/'+foldername+'/'+f, header = 0)
            readnames = df['ReadName'].tolist()
            for x in range(0, len(readnames)):
                    try:
                            tempindex = readnames[x].index('.')
                    except:
                            tempindex = readnames[x].index(':')
                    name = readnames[x][tempindex+1:-2]
                    if name not in unique_read_names and '*' not in readnames[x]:
                            readnames[x] = '*'+readnames[x]
                            removed += 1

            global rn
            rn = readnames

            df['ReadName'] = readnames
            print ('removed = '+str(removed))

            outpath = cwd+'/files/'+foldername+'/'+f
            df.to_csv(outpath, index = False)

    """
    def dataframe(f):
            cwd = os.getcwd()
            global df, dfloc, inputFilesFolder
            print ('Folder Name: ' + foldername)
            parent.logInfo('\nFolder Name: ' + foldername)
            #size = os.path.getsize(cwd+'/files/'+foldername+'/'+f)
            size = os.path.getsize(inputFilesFolder+foldername+'/'+f)
            if size > 400:
                    #df = pd.read_csv(cwd+'/files/'+foldername+'/'+f, header = 0)
                    #dfloc = cwd+'/files/'+foldername+'/'+f
                    df = pd.read_csv(inputFilesFolder+foldername+'/'+f, header = 0)
                    dfloc = inputFilesFolder+foldername+'/'+f
                    global hlc 
                    hlc = df['HostLocalCords'].tolist()
                    global vlc 
                    vlc = df['ViralLocalCords'].tolist()
                    global hrc 
                    hrc = df['HostRefCords'].tolist()
                    global vrc 
                    vrc = df['ViralRefCords'].tolist()
                    global rn 
                    rn = df['ReadName'].tolist()
                    global rl 
                    rl = df['ReadLength'].tolist()
                    global sequence 
                    sequence = df['Sequence'].tolist()
                    global chrom 
                    chrom = df['Chromosome'].tolist()
                    global gene 
                    gene = df['Gene'].tolist()
                    global ig
                    ig = df['InsideGene'].tolist()
                    global dtg 
                    dtg = df['DistanceToGene'].tolist()
                    global gd 
                    gd = df['GeneDirection'].tolist()
                    global fcs 
                    fcs = df['Focus'].tolist()
                    global go 
                    go = df['GeneObj'].tolist()
                    global fo 
                    fo = df['FocusObj'].tolist()
                    global hl 
                    hl = df['Hlength'].tolist()
                    global ho 
                    ho = df['HostOrientation'].tolist()
                    global va
                    va = df['ViralAccession'].tolist()
                    global vl 
                    vl = df['Vlength'].tolist()
                    global vo 
                    vo = df['ViralOrientation'].tolist()
                    global htm 
                    htm = df['HTM'].tolist()
                    global htma 
                    htma = df['HTMAdjusted'].tolist()
                    global vtm 
                    vtm = df['VTM'].tolist()
                    global vtma 
                    vtma = df['VTMAdjusted'].tolist()
                    global ol 
                    ol = df['Overlap'].tolist()
                    global oltm 
                    oltm = df['OverlapTM'].tolist()
                    global oltma 
                    oltma = df['OverlapTMAdjusted'].tolist()
                    global insert 
                    insert = df['Inserted'].tolist()
                    global mh 
                    mh = df['Microhomology'].tolist()
                    global hmf 
                    hmf = df['HostMapFlag'].tolist()
                    global vmf 
                    vmf = df['ViralMapFlag'].tolist()
                    global hmq 
                    hmq = df['HMapQ'].tolist()
                    global vmq 
                    vmq = df['VMapQ'].tolist()
                    global fqs 
                    fqs = df['FastqSource'].tolist()
                    global index 
                    index = df['Index'].tolist()
            else:
                    hlc = []
                    vlc = []
                    hrc = []
                    vrc = []
                    rn = []
                    rl = []
                    sequence = []
                    chrom = []
                    gene = []
                    ig = []
                    dtg = []
                    gd = []
                    fcs = []
                    go = []
                    fo = []
                    hl = []
                    ho = []
                    va = []
                    vl = []
                    vo = []
                    htm = []
                    htma = []
                    vtm = []
                    vtma = []
                    ol = []
                    oltm = []
                    oltma = []
                    insert = []
                    mh = []
                    hmf = []
                    vmf = []
                    hmq = []
                    vmq = []
                    fqs = []
                    index = []


    def getname():
            global check
            if files[x].count('.') > 1:
                    dotloc = files[x].index('.')
            else:
                    dotloc = files[x].index('_')
            global sample, filename, filename_ndx
            filename = files[x][:dotloc]
            """
            if filename == "Hep3B":
                filename_ndx = filename_ndx + 1
                filename = filename + str(filename_ndx)+ "_"
            """
            if len(hlc) == 0:
                    sample.append(filename)
            else:
                    link = viewseq.getGetSequenceFile(dfloc, files[x][:-4])
                    sample.append('=HYPERLINK("'+link+'", "'+filename+'")')
            print ('File Name: ' + str(filename))
            parent.logInfo('File Name: ' + str(filename))

            #get sample type
            global sample_names, sample_dict, sname, check, adj_name, sampletype
            fname = ''.join(ch for ch in filename if ch.isalnum()) #strip non alphanumeric characters
            if fname in sname:
                    sampletype.append(sample_dict[sample_names[sname.index(fname)]])
                    temp_name = sample_names[sname.index(fname)]
                    if temp_name in sample_names_hcc: 
                            check = True
                    """
                    if temp_name in sample_names_hcc: #if adj then HCC, store, check
                            index = sample_names_hcc.index(temp_name)
                            if sample_names_adj[index] in runfiles:
                                    check = True
                                    HCC = True
                                    adj_name = sample_names_adj[index]
                    elif temp_name in sample_names_adj: #if hcc then ADJ, store, check
                            index = sample_names_adj.index(temp_name)
                            if sample_names_hcc[index] in runfiles:
                                    check = True
                                    HCC = False
                                    adj_name = sample_names_hcc[index] 
                    """
            elif fname[-1:].upper() == 'K':
                    sampletype.append('*HCC tissue')
                    check = True
            elif fname[-1:].upper() == 'N':
                    sampletype.append('*AJC tissue')
            else:
                    sampletype.append('*')
                    print("FSS ----- force check !!!!!!!!------------")
                    check = True


    def gettotals():
            totalreads = 0
            totalhbv = 0
            section = False
            global logtxt, filename
            for x in range(0, len(files)):
                    l = len(filename)
                    if filename in files[x] and files[x].endswith('.txt') and 'log' in files[x]:
                            logtxt = files[x]
                            break
            #with open(os.getcwd()+'/files/'+foldername+'/'+logtxt) as f:
            with open(inputFilesFolder+foldername+'/'+logtxt) as f:
                    content = f.readlines()
            content = [x.strip() for x in content]
            for x in range(0, len(content)):
                    if content[x] == 'Starting Viral Alignment...':
                            section = True
                    if 'reads; of these:' in content[x] and section == True:
                            i = content[x].index('reads;')
                            num = int(content[x][:i])
                            totalreads += num
                    if '%) aligned concordantly 0 times' in content[x] and section == True:
                            i = content[x].index('(')
                            num = int(content[x][:i-1])
                            totalhbv += num
                            section = False

            global reads, HBVreads, percentHBVreads, chimericreads, consideredreads
            reads.append(totalreads)
            totalhbv = totalreads - totalhbv
            HBVreads.append(totalhbv)
            try:
                    percent = totalhbv/totalreads * 100
            except:
                    percent = 0
            percentHBVreads.append(percent)
            chimericreads.append(len(hlc))
            consideredreads.append(len(hlc))


    def gettotalmeaningful():
            global meaningfulreads, summeaningfulreads, tempseq, index, filename
            if len(hlc) > 0:
                    getcords()
                    removedup()
                    removetwosup()

                    meaningfulDICT = []
                    meaningfulDICT1 = []
                    for x in range(0, len(meaningful)):
                            meaningfulDICT.append(OrderedDict([("ReadName", rn[meaningful1[x]]), ("ReadLength", rl[meaningful1[x]]), ("Sequence", sequence[meaningful1[x]]), ("JunctionCords", (meaningful[x])), ("Start/End Cords", meaningfullength[x]), ("SupportingReads", meaningfulcount[x]), ("Chromosome", chrom[meaningful1[x]]), ("Gene", gene[meaningful1[x]]), ("InsideGene", ig[meaningful1[x]]), ("DistanceToGene", dtg[meaningful1[x]]), ("GeneDirection", gd[meaningful1[x]]), ("Focus", fcs[meaningful1[x]]), ("GeneObj", go[meaningful1[x]]), ("FocusObj", fo[meaningful1[x]]), ("HostLocalCords", hlc[meaningful1[x]]), ("Hlength", hl[meaningful1[x]]), ("HostRefCords", hrc[meaningful1[x]]), ("HostOrientation", ho[meaningful1[x]]), ("ViralAccession", va[meaningful1[x]]), ("ViralLocalCords", vlc[meaningful1[x]]), ("Vlength", vl[meaningful1[x]]), ("ViralRefCords", vrc[meaningful1[x]]), ("ViralOrientation", vo[meaningful1[x]]), ("HTM", htm[meaningful1[x]]), ("HTMAdjusted", htma[meaningful1[x]]), ("VTM", vtm[meaningful1[x]]), ("VTMAdjusted", vtma[meaningful1[x]]), ("Overlap", ol[meaningful1[x]]), ("OverlapTM", oltm[meaningful1[x]]), ("OverlapTMAdjusted", oltma[meaningful1[x]]), ("Inserted", insert[meaningful1[x]]), ("Microhomology", mh[meaningful1[x]]), ("HostMapFlag", hmf[meaningful1[x]]), ("ViralMapFlag", vmf[meaningful1[x]]), ("HMapQ", hmq[meaningful1[x]]), ("ViralMapFlag", vmf[meaningful1[x]]), ("HMapQ", hmq[meaningful1[x]]), ("VMapQ", vmq[meaningful1[x]]), ("FastqSource", fqs[meaningful1[x]]), ("Index", index[meaningful1[x]])]))			
                            meaningfulDICT1.append(OrderedDict([("ReadName", rn[meaningful1[x]]), ("ReadLength", rl[meaningful1[x]]), ("Sequence", sequence[meaningful1[x]]), ("Chromosome", chrom[meaningful1[x]]), ("Gene", gene[meaningful1[x]]), ("InsideGene", ig[meaningful1[x]]), ("DistanceToGene", dtg[meaningful1[x]]), ("GeneDirection", gd[meaningful1[x]]), ("Focus", fcs[meaningful1[x]]), ("GeneObj", go[meaningful1[x]]), ("FocusObj", fo[meaningful1[x]]), ("HostLocalCords", hlc[meaningful1[x]]), ("Hlength", hl[meaningful1[x]]), ("HostRefCords", hrc[meaningful1[x]]), ("HostOrientation", ho[meaningful1[x]]), ("ViralAccession", va[meaningful1[x]]), ("ViralLocalCords", vlc[meaningful1[x]]), ("Vlength", vl[meaningful1[x]]), ("ViralRefCords", vrc[meaningful1[x]]), ("ViralOrientation", vo[meaningful1[x]]), ("HTM", htm[meaningful1[x]]), ("HTMAdjusted", htma[meaningful1[x]]), ("VTM", vtm[meaningful1[x]]), ("VTMAdjusted", vtma[meaningful1[x]]), ("Overlap", ol[meaningful1[x]]), ("OverlapTM", oltm[meaningful1[x]]), ("OverlapTMAdjusted", oltma[meaningful1[x]]), ("Inserted", insert[meaningful1[x]]), ("Microhomology", mh[meaningful1[x]]), ("HostMapFlag", hmf[meaningful1[x]]), ("ViralMapFlag", vmf[meaningful1[x]]), ("HMapQ", hmq[meaningful1[x]]), ("VMapQ", vmq[meaningful1[x]]), ("FastqSource", fqs[meaningful1[x]]), ("Index", index[meaningful1[x]])]))
            
                    if len(meaningful) > 0:
                            output = pd.DataFrame.from_dict(meaningfulDICT1)
                            #outpath = cwd + '/files/'+foldername+'/'+ filename + 'meaningful.csv'
                            outpath = inputFilesFolder+foldername+'/'+ filename + 'meaningful.csv'
                            output.to_csv(outpath, index = False)
                            #link = viewseq.getGetSequenceFile(outpath, filename + 'meaningful.csv')
                            link = viewseq.getGetSequenceFile(outpath, filename + 'meaningful')
                            meaningfulreads.append('=HYPERLINK("'+link+'", "'+str(len(meaningfulDICT))+'")')
                            summeaningfulreads += len(meaningfulDICT)
                    else:
                            meaningfulreads.append(0)
            else:
                    meaningfulreads.append(0)
            print ('Meaningful reads: ' + str(len(meaningfulseq)))


    def blast(tempseq, x):
            result_handle = NCBIWWW.qblast(program="blastn", database="refseq_genomic", sequence=tempseq, entrez_query="GCF_000001405.37", format_type = 'text')
            #blast_result = open(os.getcwd() + '/files/' + foldername + '/' + filename + '('+ str(x+1) + ')' + '_blast.txt', "w")
            blast_result = open(inputFilesFolder + foldername + '/' + filename + '('+ str(x+1) + ')' + '_blast.txt', "w")
            blast_result.write(result_handle.read())
            print ('blast results saved to ' + inputFilesFolder + foldername + '/' + filename + '('+ str(x+1) + ')' + '_blast.txt')
            result_handle.close()
            blast_result.close()			

    def checkchromosomes(n):
            chromosomes = []
            expect = []
            chromosomes1 = []
            try:
                    #with open(os.getcwd() + '/files/' + foldername + '/' + filename + '('+ str(n+1) + ')' + '_blast.txt') as f:
                    with open(inputFilesFolder + foldername + '/' + filename + '('+ str(n+1) + ')' + '_blast.txt') as f:
                            content = f.readlines()
                            content = [c.strip() for c in content]
                            in_hit = False
                            for x in range(0, len(content)):
                                    if 'No significant similarity found.' in content[x]:
                                            return None
                                    elif '>' and 'chromosome' in content[x] and in_hit == False:
                                            i = content[x].index('chromosome')
                                            num = str(content[x][i+11:i+13])
                                            if num[1] == ' ':
                                                    num = num[0]
                                            if num.isdigit() or num == 'X' or num == 'Y':
                                                    chromosomes.append(num)
                                                    in_hit = True
                                    elif in_hit == True and 'Expect =' in content[x]:
                                            index = content[x].index('Expect')
                                            expect_value = content[x][index+9:]
                                            expect.append(float(expect_value))
                                            in_hit = False
            except:
                    if len(chromosomes) == 0:
                            return None
                    else:
                            print ('all chromosomes: ' + str(chromosomes))
                            print ('expect values: ' + str(expect))
                            for x in range(0, len(chromosomes)):
                                    if chromosomes[x] not in chromosomes1 and expect[x] <= 5.0:
                                            chromosomes1.append(chromosomes[x])
                            return chromosomes1

            for x in range(0, len(chromosomes)):
                    if chromosomes[x] not in chromosomes1 and expect[x] <= 5.0:
                            chromosomes1.append(chromosomes[x])
            return chromosomes1


    def gettotalunique():
            global sumuniquereads, uniqueDICT, index, uniquereads
            if len(hlc) > 0:
                    removesup()

                    uniqueDICT = []
                    for x in range(0, len(unique)):
                            uniqueDICT.append(OrderedDict([("ReadName", rn[unique1[x]]), ("ReadLength", rl[unique1[x]]), ("Sequence", sequence[unique1[x]]), ("Chromosome", chrom[unique1[x]]), ("Gene", gene[unique1[x]]), ("InsideGene", ig[unique1[x]]), ("DistanceToGene", dtg[unique1[x]]), ("GeneDirection", gd[unique1[x]]), ("Focus", fcs[unique1[x]]), ("GeneObj", go[unique1[x]]), ("FocusObj", fo[unique1[x]]), ("HostLocalCords", hlc[unique1[x]]), ("Hlength", hl[unique1[x]]), ("HostRefCords", hrc[unique1[x]]), ("HostOrientation", ho[unique1[x]]), ("ViralAccession", va[unique1[x]]), ("ViralLocalCords", vlc[unique1[x]]), ("Vlength", vl[unique1[x]]), ("ViralRefCords", vrc[unique1[x]]), ("ViralOrientation", vo[unique1[x]]), ("HTM", htm[unique1[x]]), ("HTMAdjusted", htma[unique1[x]]), ("VTM", vtm[unique1[x]]), ("VTMAdjusted", vtma[unique1[x]]), ("Overlap", ol[unique1[x]]), ("OverlapTM", oltm[unique1[x]]), ("OverlapTMAdjusted", oltma[unique1[x]]), ("Inserted", insert[unique1[x]]), ("Microhomology", mh[unique1[x]]), ("HostMapFlag", hmf[unique1[x]]), ("ViralMapFlag", vmf[unique1[x]]), ("HMapQ", hmq[unique1[x]]), ("VMapQ", vmq[unique1[x]]), ("FastqSource", fqs[unique1[x]]), ("Index", index[unique1[x]])]))
                    if len(unique)>0:
                            output = pd.DataFrame(uniqueDICT)
                            #outpath = cwd + '/files/' + foldername + '/' + filename + 'unique.csv'
                            outpath = inputFilesFolder + foldername + '/' + filename + 'unique.csv'
                            print("FSS ----------- "+outpath)
                            output.to_csv(outpath, index = False)
                            link = viewseq.getGetSequenceFile(outpath, filename + 'unique')
                            uniquereads.append('=HYPERLINK("'+link+'", "'+str(len(uniqueDICT))+'")')
                            sumuniquereads += len(uniqueDICT)
                    else:
                            uniquereads.append(0)

                    global uniqueTUP
                    for x in range(0, len(unique)):
                            uniqueTUP.append((unique[x], uniquecount[x], uniquelength[x], unique1[x], uniqueseq[x]))
                    uniqueTUP.sort(key=lambda tup: tup[1], reverse = True) 
            else:
                    uniquereads.append(0)
            print ('Unique Junctions: ' + str(len(uniqueseq)))

    def getsupporting():
            global suppreads, supporting, sequence, uniqueTUP, chrom
            global supportingreads, index
            if len(hlc) > 0:
                    suppcount = 0
                    suppreads1 = []
            # 	for x in range(0, len(uniqueTUP)):
            # 	suppreads.append([])
            # 	indices = [i for i, n in enumerate(ndr) if n == uniqueTUP[x][0]]
            # 	for r in indices:
            # 		suppreads[x].append(OrderedDict([("Junction Cords", ndr[r]), ("Start/End Cords", ndrlength[r]), ("Position", ndr1[r]), ("Sequence", sequence[ndr1[r]]), ("Length", str(rl[ndr1[r]])), ('Gene', gene[ndr1[r]])]))
            # 	suppreads[x].append(OrderedDict([("Junction Cords", ''), ("Start/End Cords", ''), ("Position", ''), ("Sequence", ''), ("Length", ''), ('Gene', '')]))

                    for x in range(0, len(uniqueTUP)):
                            a = uniqueTUP[x][0][0]
                            b = uniqueTUP[x][0][1]
                            indices = []
                            for i in range(0, len(ndr)):
                                    a1 = ndr[i][0]
                                    b1 = ndr[i][1]
                                    if uniqueTUP[x][4] == ndrseq[i] or reverse_complement(uniqueTUP[x][4][::-1]) == ndrseq[i]:
                                            indices.append(i)
                                    elif abs(a1 - a) <=3 and abs(b1 - b) <=3 or abs(b1 - a) <=3 and abs(a1 - b) <=3:
                                            indices.append(i)

                            tempseq = ''
                            for r in indices:
                                    suppcount += 1
                                    temp = sequence[uniqueTUP[x][3]]
                                    temp = temp[hlc1[ndr1[r]][0]: hlc1[ndr1[r]][1]]
                                    if len(temp) > len(tempseq):
                                            tempseq = temp
                            
                            # blast for all chromosomes
                            # print ('blasting for all chromosomes...')
                            # signal.signal(signal.SIGALRM, signal_handler)
                            # signal.alarm(1)   # wait time (seconds)
                            # try:
                            #     blast(tempseq, x)
                            # except Exception as msg:
                            #     print ('Timed out!')
                            # finally:
                            # 	signal.alarm(0) # disable alarm
                            
                            # chromosomes = checkchromosomes(x)
                            chromosomes = None
                            if chromosomes != None and len(chromosomes) >= 1:
                                    print ('chromsomes after blasting:' + str(chromosomes))
                                    temp = []
                                    for i in chromosomes:
                                            if i not in temp:
                                                    temp.append(i)
                                    if chrom[uniqueTUP[x][3]] not in temp:
                                            temp.append(chrom[uniqueTUP[x][3]])
                                    for i1 in indices:
                                            chrom[ndr1[i1]] = set(temp)

                                    uniqueDICT = []
                                    for i in range(0, len(unique)):
                                            uniqueDICT.append(OrderedDict([("ReadName", rn[unique1[i]]), ("ReadLength", rl[unique1[i]]), ("Sequence", sequence[unique1[i]]), ("Chromosome", chrom[unique1[i]]), ("Gene", gene[unique1[i]]), ("InsideGene", ig[unique1[i]]), ("DistanceToGene", dtg[unique1[i]]), ("GeneDirection", gd[unique1[i]]), ("Focus", fcs[unique1[i]]), ("GeneObj", go[unique1[i]]), ("FocusObj", fo[unique1[i]]), ("HostLocalCords", hlc[unique1[i]]), ("Hlength", hl[unique1[i]]), ("HostRefCords", hrc[unique1[i]]), ("HostOrientation", ho[unique1[i]]), ("ViralAccession", va[unique1[i]]), ("ViralLocalCords", vlc[unique1[i]]), ("Vlength", vl[unique1[i]]), ("ViralRefCords", vrc[unique1[i]]), ("ViralOrientation", vo[unique1[i]]), ("HTM", htm[unique1[i]]), ("HTMAdjusted", htma[unique1[i]]), ("VTM", vtm[unique1[i]]), ("VTMAdjusted", vtma[unique1[i]]), ("Overlap", ol[unique1[i]]), ("OverlapTM", oltm[unique1[i]]), ("OverlapTMAdjusted", oltma[unique1[i]]), ("Inserted", insert[unique1[i]]), ("Microhomology", mh[unique1[i]]), ("HostMapFlag", hmf[unique1[i]]), ("ViralMapFlag", vmf[unique1[i]]), ("HMapQ", hmq[unique1[i]]), ("VMapQ", vmq[unique1[i]]), ("FastqSource", fqs[unique1[i]]), ("Index", index[unique1[i]])]))
                                    output = pd.DataFrame(uniqueDICT)
                                    #outpath = cwd + '/files/' + foldername + '/' + filename + 'unique.csv'
                                    outpath = inputFilesFolder + foldername + '/' + filename + 'unique.csv'
                                    output.to_csv(outpath, index = False)

                            suppreads.append([])
                            suppreads1.append([])

                            for r in indices:
                                    #suppreads[x].append(OrderedDict([("ReadName", rn[ndr1[r]]), ("ReadLength", rl[ndr1[r]]), ("Sequence", sequence[ndr1[r]]), ("JunctionCords", ndr[r]), ("Start/End Cords", ndrlength[r]), ("Position", ndr1[r]), ("Chromosome", chrom[ndr1[r]]), ("Gene", gene[ndr1[r]]), ("InsideGene", ig[ndr1[r]]), ("DistanceToGene", dtg[ndr1[r]]), ("GeneDirection", gd[ndr1[r]]), ("Focus", fcs[ndr1[r]]), ("GeneObj", go[ndr1[r]]), ("FocusObj", fo[ndr1[r]]), ("HostLocalCords", hlc[ndr1[r]]), ("Hlength", hl[ndr1[r]]), ("HostRefCords", hrc[ndr1[r]]), ("HostOrientation", ho[ndr1[r]]), ("ViralAccession", va[ndr1[r]]), ("ViralLocalCords", vlc[ndr1[r]]), ("Vlength", vl[ndr1[r]]), ("ViralRefCords", vrc[ndr1[r]]), ("ViralOrientation", vo[ndr1[r]]), ("HTM", htm[ndr1[r]]), ("HTMAdjusted", htma[ndr1[r]]), ("VTM", vtm[ndr1[r]]), ("VTMAdjusted", vtma[ndr1[r]]), ("Overlap", ol[ndr1[r]]), ("OverlapTM", oltm[ndr1[r]]), ("OverlapTMAdjusted", oltma[ndr1[r]]), ("Inserted", insert[ndr1[r]]), ("Microhomology", mh[ndr1[r]]), ("HostMapFlag", hmf[ndr1[r]]), ("ViralMapFlag", vmf[ndr1[r]]), ("HMapQ", hmq[ndr1[r]]), ("ViralMapFlag", vmf[ndr1[r]]), ("HMapQ", hmq[ndr1[r]]), ("VMapQ", vmq[ndr1[r]]), ("FastqSource", fqs[ndr1[r]]), ("Index", index[ndr1[r]])]))
                                    suppreads1[x].append(OrderedDict([("ReadName", rn[ndr1[r]]), ("ReadLength", rl[ndr1[r]]), ("Sequence", sequence[ndr1[r]]), ("Chromosome", chrom[ndr1[r]]), ("Gene", gene[ndr1[r]]), ("InsideGene", ig[ndr1[r]]), ("DistanceToGene", dtg[ndr1[r]]), ("GeneDirection", gd[ndr1[r]]), ("Focus", fcs[ndr1[r]]), ("GeneObj", go[ndr1[r]]), ("FocusObj", fo[ndr1[r]]), ("HostLocalCords", hlc[ndr1[r]]), ("Hlength", hl[ndr1[r]]), ("HostRefCords", hrc[ndr1[r]]), ("HostOrientation", ho[ndr1[r]]), ("ViralAccession", va[ndr1[r]]), ("ViralLocalCords", vlc[ndr1[r]]), ("Vlength", vl[ndr1[r]]), ("ViralRefCords", vrc[ndr1[r]]), ("ViralOrientation", vo[ndr1[r]]), ("HTM", htm[ndr1[r]]), ("HTMAdjusted", htma[ndr1[r]]), ("VTM", vtm[ndr1[r]]), ("VTMAdjusted", vtma[ndr1[r]]), ("Overlap", ol[ndr1[r]]), ("OverlapTM", oltm[ndr1[r]]), ("OverlapTMAdjusted", oltma[ndr1[r]]), ("Inserted", insert[ndr1[r]]), ("Microhomology", mh[ndr1[r]]), ("HostMapFlag", hmf[ndr1[r]]), ("ViralMapFlag", vmf[ndr1[r]]), ("HMapQ", hmq[ndr1[r]]), ("VMapQ", vmq[ndr1[r]]), ("FastqSource", fqs[ndr1[r]]), ("Index", index[ndr1[r]])]))
                                    suppreads[x].append(OrderedDict([("Read Name", rn[ndr1[r]]), ("Junction Cords", ndr[r]), ("Start/End Cords", ndrlength[r]), ("Host Local Cords", hlc[ndr1[r]]), ("Viral Local Cords", vlc[ndr1[r]]), ("Position", ndr1[r]), ("Sequence", sequence[ndr1[r]]), ("Length", str(rl[ndr1[r]])), ('Gene', gene[ndr1[r]])]))
                            suppreads[x].append(OrderedDict([("Read Name", ''), ("Junction Cords", ''), ("Start/End Cords", ''), ("Host Local Cords", ''), ("Viral Local Cords", ''), ("Position", ''), ("Sequence", ''), ("Length", ''), ('Gene', '')]))
                            

                    if len(suppreads) > 0:
                            name = filename + 'suppreads'
                            output = pd.DataFrame(suppreads1[0])
                            for x in range(1, len(suppreads1)):
                                    output = output.append(pd.DataFrame(suppreads1[x]))
                            #outpath = cwd + '/files/'+foldername+'/'+ filename + 'suppreads.csv'
                            outpath = inputFilesFolder+foldername+'/'+ filename + 'suppreads.csv'
                            output.to_csv(outpath, index = False)
                            link = viewseq.getGetSequenceFile(outpath, filename+'suppreads')
                            supportingreads.append('=HYPERLINK("'+link+'", "'+filename+'suppreads'+'")')

                            output = pd.DataFrame(suppreads[0])
                            for x in range(1, len(suppreads)):
                                    output = output.append(pd.DataFrame(suppreads[x]))
                            supporting.append((output, name))
                            # output.to_excel(writer, filename+'suppreads', columns = headers, index = False)
                    else:
                            supportingreads.append(None)
            else:
                    supportingreads.append(None)


    def getmajorjunctionsdrivers():
            global majorjunctionreads, uniqueTUP, position, major, summajorjunctionreads, index
            global rn, rl, sequence, hlc, vlc, gene, ig, dtg, gd, fcs, go, fo, hlc, hl, hrc, ho, va
            global vlc, vl, vrc, vo, htm, htma, vtm, vtma, ol, oltm, oltma, insert, mh, hmf, vmf, hmq
            global vmf, hmq, vmq, fqs, index
            global chrom, check
            global driver
            if len(hlc) > 0:
                    uniqueTUP.sort(key=lambda tup: tup[1], reverse = True) 
                    minimum = len(meaningful)/5
                    print ('minimum = ' + str(minimum))
                    mj = []
                    mj1 = []
                    for x in range(0, len(uniqueTUP)):
                            tempreadnames = []
                            if (uniqueTUP[x][1] >= minimum):
                                    mj.append(OrderedDict([("ReadName", rn[uniqueTUP[x][3]]), ("ReadLength", rl[uniqueTUP[x][3]]), ("Sequence", sequence[uniqueTUP[x][3]]), ("JunctionCords", (uniqueTUP[x][0])), ("Start/EndCords", uniqueTUP[x][2]), ("Host Local Cords", hlc[uniqueTUP[x][3]]), ("Viral Local Cords", vlc[uniqueTUP[x][3]]), ("SupportingReads", uniqueTUP[x][1]), ("Chromosome", chrom[uniqueTUP[x][3]]), ("Gene", gene[uniqueTUP[x][3]]), ("InsideGene", ig[uniqueTUP[x][3]]), ("DistanceToGene", dtg[uniqueTUP[x][3]]), ("GeneDirection", gd[uniqueTUP[x][3]]), ("Focus", fcs[uniqueTUP[x][3]]), ("GeneObj", go[uniqueTUP[x][3]]), ("FocusObj", fo[uniqueTUP[x][3]]), ("HostLocalCords", hlc[uniqueTUP[x][3]]), ("Hlength", hl[uniqueTUP[x][3]]), ("HostRefCords", hrc[uniqueTUP[x][3]]), ("HostOrientation", ho[uniqueTUP[x][3]]), ("ViralAccession", va[uniqueTUP[x][3]]), ("ViralLocalCords", vlc[uniqueTUP[x][3]]), ("Vlength", vl[uniqueTUP[x][3]]), ("ViralRefCords", vrc[uniqueTUP[x][3]]), ("ViralOrientation", vo[uniqueTUP[x][3]]), ("HTM", htm[uniqueTUP[x][3]]), ("HTMAdjusted", htma[uniqueTUP[x][3]]), ("VTM", vtm[uniqueTUP[x][3]]), ("VTMAdjusted", vtma[uniqueTUP[x][3]]), ("Overlap", ol[uniqueTUP[x][3]]), ("OverlapTM", oltm[uniqueTUP[x][3]]), ("OverlapTMAdjusted", oltma[uniqueTUP[x][3]]), ("Inserted", insert[uniqueTUP[x][3]]), ("Microhomology", mh[uniqueTUP[x][3]]), ("HostMapFlag", hmf[uniqueTUP[x][3]]), ("ViralMapFlag", vmf[uniqueTUP[x][3]]), ("HMapQ", hmq[uniqueTUP[x][3]]), ("ViralMapFlag", vmf[uniqueTUP[x][3]]), ("HMapQ", hmq[uniqueTUP[x][3]]), ("VMapQ", vmq[uniqueTUP[x][3]]), ("FastqSource", fqs[uniqueTUP[x][3]]), ("Index", index[uniqueTUP[x][3]])]))
                                    mj1.append(OrderedDict([("ReadName", rn[uniqueTUP[x][3]]), ("ReadLength", rl[uniqueTUP[x][3]]), ("Sequence", sequence[uniqueTUP[x][3]]), ("Chromosome", chrom[uniqueTUP[x][3]]), ("Gene", gene[uniqueTUP[x][3]]), ("InsideGene", ig[uniqueTUP[x][3]]), ("DistanceToGene", dtg[uniqueTUP[x][3]]), ("GeneDirection", gd[uniqueTUP[x][3]]), ("Focus", fcs[uniqueTUP[x][3]]), ("GeneObj", go[uniqueTUP[x][3]]), ("FocusObj", fo[uniqueTUP[x][3]]), ("HostLocalCords", hlc[uniqueTUP[x][3]]), ("Hlength", hl[uniqueTUP[x][3]]), ("HostRefCords", hrc[uniqueTUP[x][3]]), ("HostOrientation", ho[uniqueTUP[x][3]]), ("ViralAccession", va[uniqueTUP[x][3]]), ("ViralLocalCords", vlc[uniqueTUP[x][3]]), ("Vlength", vl[uniqueTUP[x][3]]), ("ViralRefCords", vrc[uniqueTUP[x][3]]), ("ViralOrientation", vo[uniqueTUP[x][3]]), ("HTM", htm[uniqueTUP[x][3]]), ("HTMAdjusted", htma[uniqueTUP[x][3]]), ("VTM", vtm[uniqueTUP[x][3]]), ("VTMAdjusted", vtma[uniqueTUP[x][3]]), ("Overlap", ol[uniqueTUP[x][3]]), ("OverlapTM", oltm[uniqueTUP[x][3]]), ("OverlapTMAdjusted", oltma[uniqueTUP[x][3]]), ("Inserted", insert[uniqueTUP[x][3]]), ("Microhomology", mh[uniqueTUP[x][3]]), ("HostMapFlag", hmf[uniqueTUP[x][3]]), ("ViralMapFlag", vmf[uniqueTUP[x][3]]), ("HMapQ", hmq[uniqueTUP[x][3]]), ("VMapQ", vmq[uniqueTUP[x][3]]), ("FastqSource", fqs[uniqueTUP[x][3]]), ("Index", index[uniqueTUP[x][3]])]))
                            for s in suppreads[x]:
                                    if s['Read Name'] == '':
                                            break
                                    else:
                                            tempreadnames.append(s['Read Name'])
                            tempreadnames_str = ', '.join(tempreadnames)

                            if check:
                                    driver[position].append(OrderedDict([('Sample', filename), ("ReadNames", tempreadnames_str), ("Sequence", sequence[uniqueTUP[x][3]]), ("JunctionCords", (uniqueTUP[x][0])), ("Start/EndCords", uniqueTUP[x][2]), ("Position", uniqueTUP[x][3]), ("Chromosome", chrom[uniqueTUP[x][3]]), ("Gene", gene[uniqueTUP[x][3]]), ("SupportingReads", uniqueTUP[x][1]), ("Recurrent", None), ("RecurrencyCount", None ), ("LiterarySupport", None), ("Link", None), ('JunctionSeq', uniqueTUP[x][4])]))
                    if len(mj) > 0:
                            output = pd.DataFrame(mj1)
                            name = filename + 'majorjunctions1'
                            #outpath = cwd + '/files/'+foldername+'/'+ filename + 'majorjunctions.csv'
                            outpath = inputFilesFolder+foldername+'/'+ filename + 'majorjunctions.csv'
                            output.to_csv(outpath, index = False)
                            link = viewseq.getGetSequenceFile(outpath, filename+'majorjunctions')
                            majorjunctionreads.append('=HYPERLINK("'+link+'", "'+str(len(mj))+'")')
                            summajorjunctionreads += len(mj)

                            output = pd.DataFrame(mj)
                            name = filename + 'majorjunctions'
                            major.append((output, name))
                            # output.to_excel(writer, name, index = False)
                    else:
                            majorjunctionreads.append(0)
            else:
                    majorjunctionreads.append(0)


    def getcords(): # AND JUNCTION SEQUENCE #get all coordinates, check if human or virus comes first, and find coordinate, length, position of junction
            global ho, ho1, vo, vo1, hlc1, vlc1, rn, rc, rcseq, rclength, rc1, cumulativejunctions
            global sequence, check, check_rc, check_rcseq, check_rcname, rc_name, gene
            for i in range (0, len(ho)): #convert host orientation into tuple
                    s = ho[i]
                    x,y = s.split("/")
                    ho1.append(str(y))

            for i in range (0, len(vo)): #convert host orientation into tuple
                    s = vo[i]
                    x,y = s.split("/")
                    vo1.append(str(y))

            for i in range(0, len(hlc)): #convert hostlocalcords into tuple
                    s = hlc[i]
                    x,y = s.split(",")
                    hlc1.append((int(x[1:]), int(y[:-1])))

            for i in range(0, len(vlc)): #convert virallocalcords into tuple
                    s = vlc[i]
                    x,y = s.split(",")
                    vlc1.append((int(x[1:]), int(y[:-1])))

            for x in range(0, len(hlc1)): 
                    if hlc1[x][0] < vlc1[x][0]: #if host comes first
                            s1 = hrc[x]
                            h1, h2 = s1.split(",")
                            s2 = vrc[x]
                            v1, v2 = s2.split(",")
                            host_jcord = None # host junction cord
                            viral_jcord = None
                            host_lcord = None # viral length cord
                            viral_lcord = None
                            if ho1[x] == 'Plus': # if host is plus, use host end; use other cord for length
                                    host_jcord = int(h2[:-1])
                                    host_lcord = int(h1[1:])
                            else: # if host is minus, use host start
                                    host_jcord = int(h1[1:])
                                    host_lcord = int(h2[:-1])
                            if vo1[x] == 'Plus': # if viral is plus, use viral start
                                    viral_jcord = int(v1[1:])
                                    viral_lcord = int(v2[:-1])
                            else: # if viral is minus, use viral end
                                    viral_jcord = int(v2[:-1])
                                    viral_lcord = int(v1[1:])
                            rc.append((host_jcord, viral_jcord)) # get coordinates of junction as tuple
                            rclength.append((host_lcord, viral_lcord)) #get end coordinates of junction as tuple
                            rc1.append(x) #get position of junction in original output file
                            rc_name.append('('+ filename + ')' + rn[x])
                            #GET JUNCTION SEQUENCE
                            seq = sequence[x]
                            if hlc1[x][1] > vlc1[x][0]: #overlap
                                    overlap = hlc1[x][1] - vlc1[x][0]
                                    rcseq.append(seq[ vlc1[x][0]-6 : hlc1[x][1]+5 ])
                                    cumulativejunctions.append(((host_jcord, viral_jcord), seq[ vlc1[x][0]-6 : hlc1[x][1]+5 ], foldername))
                            else:
                                    rcseq.append(seq[ hlc1[x][1]-6 : vlc1[x][0]+5 ])
                                    cumulativejunctions.append(((host_jcord, viral_jcord), seq[ hlc1[x][1]-6 : vlc1[x][0]+5 ], foldername))


                    else: # if viral comes first
                            s1 = hrc[x]
                            h1, h2 = s1.split(",")
                            s2 = vrc[x]
                            v1, v2 = s2.split(",")
                            host_jcord = None # host junction cord
                            viral_jcord = None
                            host_lcord = None # viral length cord
                            viral_lcord = None
                            if vo1[x] == 'Plus':
                                    viral_jcord = int(v2[:-1])
                                    viral_lcord = int(v1[1:])
                            else: 
                                    viral_jcord = int(v1[1:])
                                    viral_lcord = int(v2[:-1])
                            if ho1[x] == 'Plus': 
                                    host_jcord = int(h1[1:])
                                    host_lcord = int(h2[:-1])
                            else: 
                                    host_jcord = int(h2[:-1])
                                    host_lcord = int(h1[1:])
                            rc.append((viral_jcord, host_jcord)) # get coordinates of junction as tuple
                            rclength.append((viral_lcord, host_lcord)) #get end coordinates of junction as tuple
                            rc1.append(x) #get position of junction in original output file
                            rc_name.append('('+ filename + ')' + rn[x])
                            #GET JUNCTION SEQUENCE
                            seq = sequence[x]
                            jlength = 0
                            if vlc1[x][1] > hlc1[x][0]:
                                    rcseq.append(seq[ hlc1[x][0]-6 : vlc1[x][1]+5 ])
                                    cumulativejunctions.append(((viral_jcord, host_jcord), seq[ hlc1[x][0]-6 : vlc1[x][1]+5 ], foldername))
                            else:
                                    rcseq.append(seq[ vlc1[x][1]-6 : hlc1[x][0]+5 ])
                                    cumulativejunctions.append(((viral_jcord, host_jcord), seq[ hlc1[x][0]-6 : vlc1[x][1]+5 ], foldername))

            check_rc = rc
            check_rcseq = rcseq
            check_rcname = rc_name

    def removedup():
            global rclength, ndrlength, ndr, ndr1, ndrlength, consideredreads, sequence, gene
            global ho1, vo1, hlc1, vlc1, rc, rcseq, rclength, rc_name, rc1, ndr, ndrseq, ndrlength, ndr1
            global rsize, meaningful, meaningfulseq, meaningfullength, meaningfulcount, meaningful1, uniqueTUP
            global unique, uniqueseq, uniquelength, unique1, uniquecount, suppreads, majorjunctions, index
            # for x in range(0, len(rclength)): #remove duplicate lengths
            # 	if rclength[x] not in ndrlength:
            # 		ndrlength.append(rclength[x])
            # 		ndr.append(rc[x])
            # 		ndr1.append(rc1[x])

            #df = pd.read_csv(cwd+'/files/'+foldername+'/'+ output_file, header = 0)
            df = pd.read_csv(inputFilesFolder+foldername+'/'+ output_file, header = 0)
            readnames = df['ReadName'].tolist()
            removed = []
            nd_index=[]
            print ('total junction reads: ' + str(len(rc)))
            for x in range(0, len(rclength)):
                    a = rclength[x][0]
                    b = rclength[x][1]
                    append = True

                    #remove lengths with +- 5 difference
                    for i in range(x+1, len(rclength)):
                            # if gene[x] != gene[i]: continue
                            a1 = rclength[i][0]
                            b1 = rclength[i][1]
                            if (abs(a1 - a) <= 5 and abs(b1 - b) <= 5) or (abs(b1 - a) <= 5 and abs(a1 - b) <= 5): #or '*' in readnames[x]: # check flipped and regular start/end cords 
                                    append = False
                                    if x not in removed: 
                                            removed.append(x)
                            if append==False: break

                    #remove reads with same junction and length of sequence going out from junction
                    # if append == True:
                    # 	for i in range(x+1, len(rclength)):
                    # 		if SequenceMatcher(None, rcseq[x], rcseq[i]).ratio() >= 0.90 and abs(rl[rc1[x]] - rl[rc1[i]]) <= 2:
                    # 			# A = hlc1[rc1[x]][0]
                    # 			# B = vlc1[rc1[x]][1]
                    # 			# A1 = hlc1[rc1[i]][0]
                    # 			# B1 = vlc1[rc1[i]][1]
                    # 			# if (abs(A1 - A) <= 2 and abs(B1 - B) <= 2):
                    # 			append = False
                    # 			if x not in removed: 
                    # 				removed.append(x)

                    #remove reads whose sequences are more than 90% similar
                    # if append == True:
                    # 	for i in range(x+1, len(rclength)):
                    # 		if SequenceMatcher(None, sequence[rc1[x]], sequence[rc1[i]]).ratio() >= 0.90:
                    # 			append = False
                    # 			if x not in removed: 
                    # 				removed.append(x)

                    #remove reads whose viral or human portion is less than 30 bp
                    # if abs(vlc1[x][0] - vlc1[x][1]) <= 30:
                    # 	append = False
                    # 	if x not in removed:
                    # 		removed.append(x)
                    # if abs(hlc1[x][0] - hlc1[x][1]) <= 30:
                    # 	append = False
                    # 	if x not in removed:
                    # 		removed.append(x)

                    if append == True:
                            nd_index.append(x)
                            ndrlength.append(rclength[x])
                            ndr.append(rc[x])
                            ndr1.append(rc1[x])
                            ndrseq.append(rcseq[x])
            print ('junction reads after removing duplicates: ' + str(len(ndrseq)))

            uniqueDICT = []
            for x in range(0, len(nd_index)):
                    uniqueDICT.append(OrderedDict([("ReadName", rn[nd_index[x]]), ("ReadLength", rl[nd_index[x]]),
                                ("Sequence", sequence[nd_index[x]]), ("Chromosome", chrom[nd_index[x]]),
                                ("Gene", gene[nd_index[x]]), ("InsideGene", ig[nd_index[x]]),
                                ("DistanceToGene", dtg[nd_index[x]]), ("GeneDirection", gd[nd_index[x]]),
                                ("Focus", fcs[nd_index[x]]), ("GeneObj", go[nd_index[x]]),
                                ("FocusObj", fo[nd_index[x]]), ("HostLocalCords", hlc[nd_index[x]]),
                                ("Hlength", hl[nd_index[x]]), ("HostRefCords", hrc[nd_index[x]]),
                                ("HostOrientation", ho[nd_index[x]]), ("ViralAccession", va[nd_index[x]]),
                                ("ViralLocalCords", vlc[nd_index[x]]), ("Vlength", vl[nd_index[x]]),
                                ("ViralRefCords", vrc[nd_index[x]]), ("ViralOrientation", vo[nd_index[x]]),
                                ("HTM", htm[nd_index[x]]), ("HTMAdjusted", htma[nd_index[x]]),
                                ("VTM", vtm[nd_index[x]]), ("VTMAdjusted", vtma[nd_index[x]]),
                                ("Overlap", ol[nd_index[x]]), ("OverlapTM", oltm[nd_index[x]]),
                                ("OverlapTMAdjusted", oltma[nd_index[x]]), ("Inserted", insert[nd_index[x]]),
                                ("Microhomology", mh[nd_index[x]]), ("HostMapFlag", hmf[nd_index[x]]),
                                ("ViralMapFlag", vmf[nd_index[x]]), ("HMapQ", hmq[nd_index[x]]),
                                ("VMapQ", vmq[nd_index[x]]), ("FastqSource", fqs[nd_index[x]]),
                                ("Index", index[nd_index[x]])]))
            if len(nd_index)>0:
                    output = pd.DataFrame(uniqueDICT)
                    gene_result=(output['Gene'].value_counts()).reset_index()
                    gene_result.columns=['Gene', 'Count']
                    parent.logInfo(str(gene_result))
                    outpath = inputFilesFolder + foldername + '/' + filename + 'DupRemoved.csv'
                    output.to_csv(outpath, index = False)
                    viewseq.getGetSequenceFile(outpath, filename + 'DupRemoved')

            # both = 0
            for x in removed:
                    # if '*' in readnames[x]:
                    # 	both += 1
                    if '*' not in readnames[x]:
                            readnames[x] = '*'+readnames[x]

            df['ReadName'] = readnames

            #outpath = cwd+'/files/'+foldername+'/' + output_file
            outpath = inputFilesFolder+foldername+'/' + output_file
            df.to_csv(outpath, index = False)
            # print ('both removed = ' + str(both))

            consideredreads[-1] = consideredreads[-1] - len(removed)

                                            

    def removetwosup():
            global meaningful, meaningfullength, meaningfulcount, meaningful1
            for x in range(0, len(ndr)): #remove reads with <2 supporting reads
                    # if ndr.count(ndr[x]) >= 2:
                    # 	meaningful.append(ndr[x])
                    # 	meaningfullength.append(ndrlength[x])
                    # 	meaningfulcount.append(ndr.count(ndr[x]))
                    # 	meaningful1.append(ndr1[x])
                    a = ndr[x][0]
                    b = ndr[x][1]
                    count = 0
                    for i in range(0, len(ndr)):
                            a1 = ndr[i][0]
                            b1 = ndr[i][1]
                            if abs(a1 - a) <=3 and abs(b1 - b) <=3 or abs(b1 - a) <=3 and abs(a1 - b) <=3:
                                    count += 1
                            elif SequenceMatcher(None, ndrseq[x], ndrseq[i]).ratio() >= 0.90 or SequenceMatcher(None, reverse_complement(ndrseq[x][::-1]), ndrseq[i]).ratio() >= 0.90:
                                    count += 1
                    if count >= 2:
                            meaningful.append(ndr[x])
                            meaningfullength.append(ndrlength[x])
                            meaningfulcount.append(count)
                            meaningful1.append(ndr1[x])
                            meaningfulseq.append(ndrseq[x])


    def removesup():
            global unique, uniquecount, uniquelength, unique1
            for x in range(0, len(meaningful)): #remove supporting reads
                    a = meaningful[x][0]
                    b = meaningful[x][1]
                    append = True	
                    for i in range(x+1, len(meaningful)):
                            a1 = meaningful[i][0]
                            b1 = meaningful[i][1]
                            if (abs(a1 - a) <=3 and abs(b1 - b) <=3) or (abs(b1 - a) <=3 and abs(a1 - b) <=3):
                                    append = False
                            elif SequenceMatcher(None, meaningfulseq[x], meaningfulseq[i]).ratio() >= 0.90 or SequenceMatcher(None, reverse_complement(meaningfulseq[x][::-1]), meaningfulseq[i]).ratio() >= 0.90:
                                    append = False
                    if append == True:
                            unique.append(meaningful[x])
                            uniquecount.append(meaningfulcount[x])
                            uniquelength.append(meaningfullength[x])
                            unique1.append(meaningful1[x])
                            uniqueseq.append(meaningfulseq[x])


    def checkadjnontumor():
            return
            """
            global check_rc, check_rcseq, driver
            #determine risk, +1 to 3rd "Risk" in driver if not found in adjacent nontumor
            print ('checking adj nontumor...')
            for x in driver[position]:
                    if HCC:
                            if x['Sample'] == filename:
                                    print ('checking: ' + adj_name)
                                    a = x['JunctionCords'][0]
                                    b = x['JunctionCords'][1]
                                    print ('driver cords = ' + str(x['JunctionCords']))
                                    print ('driver seq = ' + x['JunctionSeq'])
                                    junctionseq = x['JunctionSeq']
                                    count = 0
                                    readnames = []
                                    for i in range(0, len(check_rc)):
                                            a1 = check_rc[i][0]
                                            b1 = check_rc[i][1]
                                            if abs(a1 - a) <=3 and abs(b1 - b) <=3 or abs(b1 - a) <=3 and abs(a1 - b) <=3:
                                                    count += 1
                                                    readnames.append(check_rcname[i])
                                            elif SequenceMatcher(None, junctionseq, check_rcseq[i]).ratio() >= 0.90 or SequenceMatcher(None, reverse_complement(junctionseq[::-1]), check_rcseq[i]).ratio() >= 0.90:
                                                    count += 1
                                                    readnames.append(check_rcname[i])
                                    if count >= 2:
                                            x["AdjacentNontumor"] = "X"
                                            x["AdjNontumorReads"] = ', '.join(map(str, readnames))
                                            x["Risk"] += 1
                    else:
                            if x['Sample'] == adj_name:
                                    print ('checking: ' + filename)
                                    a = x['JunctionCords'][0]
                                    b = x['JunctionCords'][1]
                                    print ('driver cords = ' + str(x['JunctionCords']))
                                    print ('driver seq = ' + x['JunctionSeq'])
                                    junctionseq = x['JunctionSeq']
                                    count = 0
                                    readnames = []
                                    for i in range(0, len(rc)):
                                            a1 = rc[i][0]
                                            b1 = rc[i][1]
                                            if abs(a1 - a) <=3 and abs(b1 - b) <=3 or abs(b1 - a) <=3 and abs(a1 - b) <=3:
                                                    count += 1
                                                    readnames.append(rc_name[i])
                                            elif SequenceMatcher(None, junctionseq, rcseq[i]).ratio() >= 0.90 or SequenceMatcher(None, reverse_complement(junctionseq[::-1]), rcseq[i]).ratio() >= 0.90:
                                                    count += 1
                                                    readnames.append(rc_name[i])
                                    if count >= 2:
                                            x["AdjacentNontumor"] = "X"
                                            x["AdjNontumorReads"] = ', '.join(map(str, readnames))
                                            x["Risk"] += 1
            """


    def checkpubmed():
            global searched, driver, position
            print ('data mining in pubmed...') 
            parent.logInfo ('data mining in pubmed...') 
            for x in range(0, len(driver[position])):
                    g = str(driver[position][x]["Gene"])
                    if g == "nan": continue
                    url = "https://www.ncbi.nlm.nih.gov/pubmed?term=("+g+"%20AND%20Cancer)%20OR%20("+g+"%20AND%20Tumor)"
                    if searched.count((g,"X")) > 0:
                            driver[position][x]["LiterarySupport"] = "X"
                            driver[position][x]["Link"] = url
                    elif searched.count((g, None)) > 0:
                            driver[position][x]["LiterarySupport"] = None
                    elif str(g) != 'nan':
                            try:
                                    page = urllib.request.urlopen(url)
                                    data = str(page.read())

                                    if data.count("No documents match your search terms") == 0:
                    # if g != 'NaN' and g != 'None':
                                            driver[position][x]["LiterarySupport"] = "X"
                                            driver[position][x]["Link"] = url
                                            searched.append((g,"X"))
                            except urllib.error.URLError:
                                    print("\nURL access denied!\nIt could be a network security issue.\n")
                                    pass
                            except:
                                    pass
                            # else:
                            # 	searched.append((g, None))


    def checkrecurrency():
            global cumulativejunctions, driver, position
            if os.path.isfile('DriverGeneDatabase.xlsx'): #if driver database does not exist
                    book = load_workbook('DriverGeneDatabase.xlsx')
                    writer = ExcelWriter('DriverGeneDatabase.xlsx', engine = 'openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                    driver_list = []
                    samples = []
                    for cell in range(1, len(writer.sheets['Driver Genes']['A'])):
                            driver_list.append(writer.sheets['Driver Genes']['A'][cell].value)
                            # sample1 = writer.sheets['Driver Genes']['G'][cell].value
                            # if str(sample1) not in samples:
                            # 	samples.append(sample1)

                    for i in range(0, len(driver[position])):
                            gene = driver[position][i]["Gene"] 
                            if gene in driver_list:
                                    driver[position][i]["Recurrent"] = "YES"


    def reset():
            global check, check_rc, check_rcseq, ho1, vo1, hlc1, vlc1, rc, rcseq, rclength, rc_name, rc1, ndr, ndrseq, ndrlength, ndr1, rsize, meaningful, meaningfulseq, meaningfullength, meaningfulcount, meaningful1, uniqueTUP, unique, uniqueseq, uniquelength, unique1, uniquecount, suppreads, majorjunctions
            check = False
            check_rc = []
            check_rcseq = []
            hlc1 = []
            vlc1 = []
            ho1= []
            vo1 = []
            rc = []
            rclength = []
            rc1 = []
            ndr = [] #noduplicatereads
            ndrlength = []
            ndr1 = []
            meaningful = [] #meaningfulreads
            meaningfullength = []
            meaningfulcount = []
            meaningful1 = []
            unique = [] #no supporting reads
            uniquelength = []
            unique1 = []
            uniquecount = []
            suppreads = []
            majorjunctions = []
            uniqueTUP = []
            rcseq = []
            ndrseq = []
            meaningfulseq = []
            uniqueseq = []
            rc_name = []

    def in_dictlist(gene):
        global drivergenes
        for x in range(0, len(drivergenes)):
            if drivergenes[x]['Gene'] == gene:
                return x
        return -1

    # -----------------------------------------------------------------------------------------------
    # -----------------------------------------------------------------------------------------------

    cwd = os.getcwd()
    sub = 0
    sub1 = 0
    rows = 2

    print ('\n')

    global driver, check

    print("FSS ----------------------------------------------------- ")
    all_summaries_folder = cwd.replace("\\","/")+"/_summaries"
    print("folder = "+all_summaries_folder)
    if not os.path.exists(all_summaries_folder):
            os.makedirs(all_summaries_folder)
    summary_folder = all_summaries_folder + "/files"
    if not os.path.exists(summary_folder):
            os.makedirs(summary_folder)
    else:
            c_t = datetime.datetime.now()
            c_t_str=str(c_t.year)+"_"+str(c_t.month)+"_"+str(c_t.day)+"_"
            c_t_str=c_t_str + str(c_t.hour)+"_"+str(c_t.minute)+"_"+str(c_t.second)
            summary_folder_renamed = all_summaries_folder + "/_files_renamed_"+c_t_str
            os.rename(summary_folder, summary_folder_renamed)
            os.makedirs(summary_folder)
    for rr in range(0, len(folders)):
            if folders[rr].startswith('.'): # hidden files
                    continue
            if os.path.isfile(inputFilesFolder + folders[rr]): # it's a file
                    continue
            foldername = folders[rr]
            files = os.listdir(inputFilesFolder + foldername)
            for x in range(0, len(files)): #get information for summary sheet
                    if files[x].startswith('.'):
                            continue
                    dirname, basename = os.path.split(files[x])
                    if ('_final' in files[x]) and ('.csv' in files[x]):
                            copyfile(inputFilesFolder+foldername+"/"+files[x], summary_folder+"/"+basename)
                    if ('_log' in files[x]) and ('.txt' in files[x]):
                            copyfile(inputFilesFolder+foldername+"/"+files[x], summary_folder+"/"+basename)
    inputFilesFolder = all_summaries_folder+"/"
    folders = os.listdir(inputFilesFolder)
    print("FSS ----------------------------------------------------- ")

    for r in range(0, len(folders)): #iterate through folder with output .csv and log .txt files
            viewseq = newViewSequence_1_11.viewSequence() 

            if (folders[r].startswith('.')) or (folders[r].startswith('_')): #account for hidden files in folder
                    sub += 1
            elif os.path.isfile(inputFilesFolder + folders[r]): #it is a file, not folder
                    sub += 1
            else:
                    foldername = folders[r]
                    #files = os.listdir(cwd + '/files/' + foldername)
                    files = os.listdir(inputFilesFolder + foldername)
                    driver.append([])
                    position = r - sub
                    global sampletype, sample, reads, HBVreads, percentHBVreads
                    global chimericreads, consideredreads, meaningfulreads, uniquereads
                    global majorjunctionreads, summeaningfulreads, sumuniquereads
                    global summajorjunctionreads, supportingreads, supporting
                    global major

                    for x in range(0, len(files)): #get information for summary sheet
                            if not files[x].startswith('.') and 'final' in files[x] and '.csv' in files[x]:
                                    start1 = timer()
                                    rows += 1
                                    output_file = files[x]
                                    dataframe(files[x])
                                    check = False
                                    getname()
                                    # for i in files:
                                    # 	if '.sam' in i and not i.startswith('[') and filename in i and 'hostAlign' in i:
                                    # 		markdups(files[x], i)
                                    # 		break
                                    gettotals()
                                    print ('identifying major junctions...')
                                    parent.logInfo ('identifying major junctions...')
                                    gettotalmeaningful()
                                    gettotalunique()
                                    getsupporting()
                                    getmajorjunctionsdrivers()
                                    runfiles.append(filename)
                                    for g in gene:
                                            g = str(g)
                                            cumulativegenes.append(g)
                                    checkrecurrency()
                                    if check:
                                        print ('identifying drivers...')
                                        parent.logInfo('identifying drivers...')
                                    reset()
                                    print (timer() - start1)
                                    print ()

                    sampletype.append('')
                    sample.append('Total') #get totals
                    reads.append(sum(reads))
                    HBVreads.append(sum(HBVreads))
                    try:
                            percentHBVreads.append(sum(HBVreads)/sum(reads)*100)
                    except:
                            percentHBVreads.append('-')
                    chimericreads.append(sum(chimericreads))
                    consideredreads.append(sum(consideredreads))
                    meaningfulreads.append(summeaningfulreads)
                    uniquereads.append(sumuniquereads)
                    # supportingreads.append(sum(supportingreads))
                    majorjunctionreads.append(summajorjunctionreads)

                    finalsummary = [] #list of lists
                    finalsummary.append(sampletype)
                    finalsummary.append(sample)
                    finalsummary.append(reads)
                    finalsummary.append(HBVreads)
                    finalsummary.append(percentHBVreads)
                    finalsummary.append(chimericreads)
                    finalsummary.append(consideredreads)
                    finalsummary.append(meaningfulreads)
                    finalsummary.append(uniquereads)
                    finalsummary.append(supportingreads)
                    finalsummary.append(majorjunctionreads)

                    headers = ['Type', 'Sample', 'NGS Reads', 'HBV Reads', 'Percent HBV Reads', 'Chimeric Reads', 'Considered Reads', 'Meaningful Reads', 'Unique Reads', 'Supporting Reads', 'Major Junctions']
                    summary = pd.DataFrame(finalsummary) #convert to pandas dataframe
                    summary = summary.transpose()
                    summary.columns = headers

                    summeaningfulreads = 0 #reset final summary vars
                    sumuniquereads = 0
                    summajorjunctionreads = 0
                    sampletype = []
                    sample = []
                    reads = []
                    HBVreads = []
                    percentHBVreads = []
                    chimericreads = []
                    consideredreads = []
                    meaningfulreads = []
                    uniquereads = []
                    supportingreads = []
                    majorjunctionreads = []

                    checkpubmed() #check pubmed

                    summaryname = foldername
                    if '_' in foldername:
                        summaryname = foldername[:foldername.index('_')]
                    #outpath = cwd + '/files/' + foldername + '/_' +summaryname+'_Summary.xlsx'
                    outpath = inputFilesFolder + foldername + '/_' +summaryname+'_Summary.xlsx'
                    writer = ExcelWriter(outpath)
                    summary.to_excel(writer, 'Summary', index = False)
                    
                    headers = ['Sample', 'Read Name', 'Read Length', 'Sequence', 'Junction Cords', 'Start/End Cords', 'Host Local Cords', 'Viral Local Cords', 'Position', 'Gene', 'Supporting Reads', 'Adjcnt Nontumor', 'Adj Nontumor Reads', 'Lit Support', 'Link', 'Recurrency', 'Recurrency Count', 'Risk']
                    for x in driver[position]:
                            del x['JunctionSeq']
                    driversummary = pd.DataFrame(driver[position])
                    # driversummary.columns = headers
                    driversummary.to_excel(writer, 'Driver', index = False)

                    x = 0
                    for x in range(0, len(supporting)):
                            output = supporting[x][0]
                            name = supporting[x][1]
                            headers = ["Read Name", "Junction Cords", "Start/End Cords", "Host Local Cords", "Viral Local Cords", "Position", "Sequence", "Length", "Gene"]
                            output.to_excel(writer, name, columns = headers, index = False)

                            try:
                                    output1 = major[x][0]
                                    name1 = major[x][1]
                                    output1.to_excel(writer, filename, index = False)
                            except:
                                    pass

                    writer.save()
                    """
                    print ('Summary file saved as: ' + str(outpath) + '\n')
                    parent.logInfo ('Summary file saved as: ' + str(outpath) + '\n')
                    """
                    wb = load_workbook(outpath)
                    ws = wb.get_sheet_by_name('Summary')
                    letters = ['B', 'H', 'I', 'J', 'K']
                    for x in range(2, rows): #alignment, hyperlink formatting
                            for l in letters:
                                    s = l + str(x)
                                    cell = ws[str(s)]
                                    if cell.value != 0 and 'HYPERLINK' in str(cell.value):
                                            cell.font = Font(color = "FF0000FF", underline = "single")
                                            cell.alignment = Alignment(horizontal = "right")
                                    else:
                                            cell.alignment = Alignment(horizontal = "right")
                    wb.save(outpath)
                    supporting = []
                    major = []
                    rows = 2
                    global drivergenes

                    # Add to Driver Gene Database
                    if not os.path.isfile('DriverGeneDatabase.xlsx'): #if driver database does not exist
                            pass
                            """
                            print ('\nCreating Driver Database file... \n')
                            writer = ExcelWriter('DriverGeneDatabase.xlsx')
                            for pos in driver: #for every sample in driver
                                    for x in pos:
                                            if not x['Gene'] != x['Gene']: #check if 'Gene' is NaN
                                                    if in_dictlist(x['Gene']) == -1: #if gene is not in drivergenes list
                                                            print (x)
                                                            drivergenes.append(OrderedDict([('Gene', x['Gene']), ('Supporting Reads', x['SupportingReads']), ('Risk', x['Risk']), ('PubMed Link', x['Link']), ('Reads', [x['ReadNames']])]))
                                                    else:
                                                            index = in_dictlist(x['Gene']) #get index of gene in list and add to supporting reads, read names
                                                            drivergenes[index]['Supporting Reads'] += x['SupportingReads']
                                                            letters = [l for l in x['ReadNames'] if l!= '[' and l!= ']' and l != "'" and l!= ',']
                                                            readname= ''.join(map(str, letters))
                                                            drivergenes[index]['Reads'].append(readname)

                            for index in drivergenes: #formatting of read names
                                    letters = [x for x in str(index['Reads']) if x != '[' and x!= ']' and x != "'" and x!= ',']
                                    index['Reads'] = ''.join(map(str, letters))
                                    index['Reads'] = index['Reads'].replace(' ', ', ')
                            print (drivergenes)
                            drivergenes = sorted(drivergenes, key = lambda k: k['Supporting Reads'], reverse = True)
                            drivergenes_df = pd.DataFrame(drivergenes)
                            drivergenes_df = drivergenes_df[['Gene','Supporting Reads','Risk','PubMed Link','Reads']]
                            drivergenes_df.to_excel(writer, 'Driver Genes', index = False)
                            writer.save()
                            """
                    else: #if driver gene database already exists
                            print ('\nAppending results to Driver Database... \n')
                            book = load_workbook('DriverGeneDatabase.xlsx')
                            writer = ExcelWriter('DriverGeneDatabase.xlsx', engine = 'openpyxl')
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # Convert sheet to dataframe
                            found_driver_genes = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
                            for c in range(0, len(found_driver_genes)):
                                    templist = []
                                    for col in writer.sheets['Driver Genes'][found_driver_genes[c]]:
                                            templist.append(col.value)
                                    found_driver_genes[c] = templist[1:]
                            # headers = ['Gene', 'Supporing Reads', 'Reads']

                            # get new driver genes
                            drivergenes = []
                            
                            for pos in driver:
                                    for x in pos:
                                            if not x['Gene'] != x['Gene']:
                                                    if in_dictlist(x['Gene']) == -1:
                                                            drivergenes.append(OrderedDict([('Gene', x['Gene']), ('Supporting Reads', x['SupportingReads']), ('Recurrent', x['Recurrent']), ('RecurrencyCount', x['RecurrencyCount']), ('PubMed Link', x['Link']), ('Samples', [x['Sample']]), ('Reads', [x['ReadNames']])]))
                                                            print (x['Gene'], x['Recurrent'])
                                                    else:
                                                            index = in_dictlist(x['Gene'])
                                                            try:
                                                                    comma = str(x['ReadNames']).index(',')
                                                            except:
                                                                    comma = len(x['ReadNames'])
                                                            append = True
                                                            if str(x['ReadNames'])[1:comma] in drivergenes[index]['Reads']:
                                                                    append = False
                                                            if append:
                                                                    drivergenes[index]['Supporting Reads'] += x['SupportingReads']
                                                                    if str(x['Sample']) not in str(drivergenes[index]['Samples']) and 'None' not in str(x['Sample']):
                                                                            drivergenes[index]['Samples'].append(x['Sample'])
                                                                    drivergenes[index]['RecurrencyCount'] = len(drivergenes[index]['Samples'])


                                                            letters = [l for l in x['ReadNames'] if l!= '[' and l!= ']' and l != "'" and l!= ',']
                                                            readname= ''.join(map(str, letters))
                                                            # print (readname)
                                                            # print (drivergenes[index]['Reads'])
                                                            # print (type(drivergenes[index]['Reads']))
                                                            drivergenes[index]['Reads'].append(str(readname))

                                                            # print (readname)
                                                            # print (drivergenes[index]['Reads'])
                                                            # print (type(drivergenes[index]['Reads']))

                            for loc in range(0, len(found_driver_genes[0])):
                                    if in_dictlist(found_driver_genes[0][loc]) > -1:
                                            index = in_dictlist(found_driver_genes[0][loc])
                                            append = True

                                            if found_driver_genes[6][loc] != None:
                                                    try:
                                                            comma = str(found_driver_genes[6][loc]).index(',')
                                                    except:
                                                            comma = len(found_driver_genes[6][loc])

                                                    if found_driver_genes[6][loc][1:comma] in str(drivergenes[index]['Reads']):
                                                            append = False
                                                    # print ('found, ' + str(drivergenes[index]['Supporting Reads']))
                                                    # try:
                                                    # 	letters = [x for x in str(found_driver_genes[7][loc]) if x != '[' and x!= ']' and x != "'" and x!= ',']
                                                    # except:
                                                    # 	pass
                                                    # readname= ''.join(map(str, letters))
                                                    # drivergenes[index]['Reads'].append(readname)

                                            if append: 
                                                    if found_driver_genes[6][loc] != None:
                                                            drivergenes[index]['Supporting Reads'] += found_driver_genes[1][loc]
                                                    try:
                                                            letters = [x for x in str(found_driver_genes[6][loc]) if x != '[' and x!= ']' and x != "'" and x!= ',']
                                                    except:
                                                            pass
                                                    readname= ''.join(map(str, letters))
                                                    drivergenes[index]['Reads'].append(readname)
                                                    if found_driver_genes[5][loc] != None and 'None' not in found_driver_genes[5][loc] and found_driver_genes[5][loc] not in drivergenes[index]['Samples'] and found_driver_genes[5][loc] not in str(drivergenes[index]['Samples']):
                                                            print (found_driver_genes[5][loc])
                                                            drivergenes[index]['Samples'].append(found_driver_genes[5][loc])


                                    else:
                                            drivergenes.append(OrderedDict([('Gene', found_driver_genes[0][loc]), ('Supporting Reads', found_driver_genes[1][loc]), ('Recurrent', found_driver_genes[2][loc]), ('RecurrencyCount', found_driver_genes[3][loc]), ('PubMed Link', found_driver_genes[4][loc]), ('Samples', [found_driver_genes[5][loc]]), ('Reads', [found_driver_genes[6][loc]])]))

                            for index in drivergenes:
                                    if 'None' not in str(index['Samples']):
                                            index['RecurrencyCount'] = len(index['Samples'])
                                    if index['RecurrencyCount'] > 1 and 'None' not in index['Samples']:
                                            index['Recurrent'] = 'YES'
                                    letters = [x for x in str(index['Samples']) if x != '[' and x!= ']' and x != "'" and x!= ',']
                                    index['Samples'] = ''.join(map(str, letters))
                                    index['Samples'] = index['Samples'].replace(' ', ', ')
                                    del index['Reads']


                            if r == len(folders) - 1:
                                    print ('POP:')
                                    pop = []
                                    for dg in range(0, len(drivergenes)):
                                            if drivergenes[dg]['Recurrent'] != 'YES' and drivergenes[dg]['PubMed Link'] == None:
                                                    pop.append(dg)
                                    for p in reversed(pop):
                                            print (drivergenes[p]["Gene"])
                                            drivergenes.pop(p)

                            drivergenes = sorted(drivergenes, key = lambda k: k['RecurrencyCount'], reverse = True)
                            # print (drivergenes)
                            drivergenes_df = pd.DataFrame(drivergenes)
                            drivergenes_df.to_excel(writer, 'Driver Genes', index = False)
                            writer.save()
                            drivergenes = []


    cwd = os.getcwd()

    driver=[]

    # get current date & time in string
    c_t = datetime.datetime.now()
    c_t_str=str(c_t.year)+"_"+str(c_t.month)+"_"+str(c_t.day)+"_"
    c_t_str=c_t_str + str(c_t.hour)+"_"+str(c_t.minute)+"_"+str(c_t.second)
    database_filename = "DriverGeneDatabase_"+c_t_str+".xlsx"
    copyfile("DriverGeneDatabase.xlsx", database_filename)
    new_summary_folder = all_summaries_folder + "/_files__"+c_t_str
    os.rename(summary_folder, new_summary_folder)
    dirname, basename = os.path.split(outpath)
    parent.logInfo("-----------------------------------------")
    parent.logInfo("database created:\n  "+database_filename)
    parent.logInfo("summary saved as:\n  "+new_summary_folder+"/"+basename)
    parent.logInfo("-----------------------------------------\n")

    

    time_used = timer() - start
    print(str(time_used))

    data="Complete, time used = "+ str(time_used)+"\n"

    parent.logInfo(data)
    parent.enableOptions()

    # sys.stdout = old_stdout
    # log_file.close()

    #-----------------------------------------------------------------------------------------------------------------
"""
process(5)
"""

