#import pysam
import re
import os
import os.path
import sys
from ast import literal_eval
from Bio import SeqIO
import pandas as pd
import numpy as np
from collections import OrderedDict
from collections import Counter
from pandas import ExcelWriter
from pandas import ExcelFile
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
import urllib.request, urllib.parse, urllib.error
from timeit import default_timer as timer
import time
import platform
import datetime
import shutil


def checkRequirements(parent, filesFolder):
    status = True
    if not os.path.isfile("HBV_Mutations.xlsx"):
        parent.logInfo("Error: file HBV_Mutations.xlsx is not found")
        status = False
    if (filesFolder == None):
        if os.path.exists("viralsamfiles"):
            if os.path.isfile("viralsamfiles"):
                  parent.logInfo("Error: folder 'viralsamfiles' is not found")
                  parent.logInfo("       create 'viralsamfiles' folder and put testfiles into 'viralsamfiles' folder")
                  status = False
        else:
            parent.logInfo("Error: folder 'viralsamfiles' is not found")
            parent.logInfo("       create 'viralsamfiles' folder and put testfiles into 'viralsamfiles' folder")
            status = False
    return status
    
def process(parent, filesFolder=None, HBV_reference=None):

    debug_count = 0
    
    def logInfo(message):
        if (parent != None):
            parent.logInfo(message)
        else:
            print(message)
        return

    def showStatus(message):
        if (parent != None):
            parent.showStatus(message)
        else:
            print(message)
        return

    start = timer()

    logInfo("start the process...")
    
    old_stdout = sys.stdout
    log_file=None
    """
    log_file = open(os.getcwd() + "/viralsamfiles/hbvmutation_log.txt","w")
    sys.stdout = log_file
    """

    if (filesFolder == None):
        print ('Identify HBV Mutations from folder viralsamfiles\n')
        selectedFolder = 'viralsamfiles'
        folders = os.listdir('viralsamfiles')
    else:
        print ('Identify HBV Mutations from folder ', filesFolder, "\n")
        selectedFolder = filesFolder
        folders=()
        for dirpath, dirnames, files in os.walk(selectedFolder):
            for name in files:
                if ('viralAlign.sam' in name) and (not '[' in name):
                    viral_samfile = os.path.join(dirpath, name).replace("\\", "/")
                    folders = folders+(viral_samfile,)

    filename = ''
    samplename = ''

    def in_dictlist(mut_location):
        for x in range(0, len(orderedmutations)):
            if orderedmutations[x]['Location'] == mut_location:
                return x
        return -1

    def convertstring(mylist):
            newstring = []
            temp = ''
            append = False
            mylist = str(mylist)
            for i in range(0, len(mylist)):
                    l = mylist[i]
                    if l.isalnum() or l == '_':
                            temp += l
                    elif (l == ',' and mylist[i-1].isalnum()) or (l == "'" and mylist[i-1].isalnum()):
                            newstring.append(temp)
                            temp = ''
            return newstring

    def find_mutant_max(x, mutant_str, debug=False):
        max_type=""
        max_count=-1
        mutant_str=mutant_str.replace("'", "").replace("{", "").replace("}", "").replace(" ", "")
        items=mutant_str.split(",")
        if (debug):
            print("DEBUG", x, mutant_str, items)
        for i in items:
            item_type, item_count = i.split(":")
            if item_type != x:
                if (debug):
                    print("DEBUG", x, item_type, item_count)
                if int(item_count) > int(max_count):
                    max_count = item_count
                    max_type = item_type
        return (max_type, max_count)

    if os.path.isfile('MutationDatabase.xlsx'):
        shutil.move('MutationDatabase.xlsx', 'MutationDatabase_old.xlsx')

    df = pd.read_excel('HBV_Mutations.xlsx', header = 0)
    IUPAC_amino = {'I': ['ATT', 'ATC', 'ATA'], 'L': ['CTT', 'CTC', 'CTA', 'CTG', 'TTA', 'TTG'], 'V': ['GTT', 'GTC', 'GTA', 'GTG'], 'F': ['TTT', 'TTC'], 'M': ['ATG'], 'C': ['TGT', 'TGC'], 'A': ['GCT', 'GCC', 'GCA', 'GCG'], 'G': ['GGT', 'GGC', 'GGA', 'GGG'], 'P': ['CCT', 'CCC', 'CCA', 'CCG'], 'T': ['ACT', 'ACC', 'ACA', 'ACG'], 'S': ['TCT', 'TCC', 'TCA', 'TCG', 'AGT', 'AGC'], 'Y': ['TAT', 'TAC'], 'W': ['TGG'], 'Q': ['CAA', 'CAG'], 'N': ['AAT', 'AAC'], 'H': ['CAT', 'CAC'], 'E': ['GAA', 'GAG'], 'D': ['GAT', 'GAC'], 'K': ['AAA', 'AAG'], 'R': ['CGT', 'CGC', 'CGA', 'CGG', 'AGA', 'AGG'], 'STOP': ['TAA', 'TAG', 'TGA']}
    IUPAC_nt = {'A': ['A'], 'C': ['C'], 'G': ['G'], 'T': ['T'], 'U': ['U'], 'R': ['A', 'G'], 'Y':['C', 'T'], 'S': ['G', 'C'], 'W': ['A', 'T'], 'K':['G', 'T'], 'M':['A', 'C'], 'B': ['C', 'G', 'T'], 'D': ['A', 'G', 'T'], 'H': ['A', 'C', 'T'], 'V': ['A', 'C', 'G'], 'N':['A', 'C', 'G', 'T'], '.':['gap'], '-':['gap']}
    # knownmutations = [('A1762T/G1764A', 'HCC', ''), ('G1896A', 'HCC'), ('G1899A', 'HCC'), ('Y1753V', 'HCC'), ('C1653T', 'HCC'), ('N1386M', 'HCC, X ORF'), ('N1485T', 'HCC, X ORF'), ('N1499B', 'HCC, X ORF'), ('G1613A', 'HCC, X ORF'), ('A1727G', 'HCC, X ORF'), ('N1757A/N1764Y/N1766G', 'HCC, X ORF'), ('C1773T', 'HCC, X ORF'), ('N1766T/N1768A', 'HCC, X ORF'), ('T1674S', 'HCC'), ('T1909C', 'HCC'), ('T1761C', 'HCC'), ('T1741C', 'HCC'), ('N1719N', 'HCC, CHB, Cirrhosis'), ('N1913N', 'HCC, CHB, Cirrhosis'), ('N1846N', 'HCC, CHB, Cirrhosis'), ('C1673Y', 'HCC, Cirrhosis'), ('C1673T/A1679G', 'HCC')]
    knownmutations = df['Mutation'].tolist()[0:29]
    function = df['Functional Significance'].tolist()[0:29]
    reflinks = df['Reference(s)'].tolist()[0:29]
    foundmutations = []
    orderedmutations = []
    results = []

    reference = ''
    readcords = tuple()
    alignedseq = ''
    referenceseq = ''
    genotype = ''

    if HBV_reference == None:
        HBV_data_dict = SeqIO.to_dict(SeqIO.parse(open('HBV_complete.fa'), 'fasta'))
    else:
        HBV_data_dict = SeqIO.to_dict(SeqIO.parse(open(HBV_reference), 'fasta'))
    
    t1 = time.time()
    t_m=0
    t_d=0
    platformTest=platform.system()

    """
    print("FSS files to be processed ---------------------")
    for r in range(0, len(folders)):
            if not folders[r].startswith('.') and not folders[r].startswith('[') and folders[r].endswith('.sam'):
                    filename = folders[r]
                    print (filename)
    print("FSS -------------------------------------------\n")
    """

    # FSS -------------------------
    # initial settings
    read_ref_length = 4000
    read_locations=[0]*(read_ref_length+1) 
    read_locations_counts=[0]*(read_ref_length+1)
    read__types=[" "]*(read_ref_length+1)
    read_max_type=[" "]*(read_ref_length+1)
    read_max_count=[0]*(read_ref_length+1)
    add_read_locations=True
    add_mut_count=1
    first_reference = None
                    
    
    for r in range(0, len(folders)):
            if not folders[r].startswith('.') and not folders[r].startswith('[') and folders[r].endswith('.sam'):
                    filename = folders[r]
                    print (filename)
                    debug_count = 0
                    #samtools view -hS -F4 sample.sam > sample.mapped.sam
                    if (filesFolder == None):
                        samfile_in = os.getcwd() + '/viralsamfiles/' + filename
                        samfile_in = samfile_in.replace("\\", "/")
                    else:
                        samfile_in = filename
                    samfile_in_orig = samfile_in
                    print ('processing ' + samfile_in + ' ... \n')
                    logInfo ('\n\nprocessing ' + samfile_in+ ' ... \n')
                    if (filesFolder == None):
                        samfile_out = os.getcwd() + '/viralsamfiles/[new]' + filename
                        samfile_out = samfile_out.replace("\\", "/")
                    else:
                        dirpath, name = os.path.split(filename)
                        samfile_out = os.path.join(dirpath, '[new]'+name)
                    
                    # get mapped reads
                    # if header is needed, use option '-hS' instead of '-S'
                    # https://github.com/alvaralmstedt/Tutorials/wiki/Separating-mapped-and-unmapped-reads-from-libraries
                    command0 = "samtools view -hS -F4 "+samfile_in+" > " + samfile_out
                    os.system(command0)

                    # convert to bam format
                    samfile_in=samfile_out
                    samfile_out=samfile_in.replace(".sam", ".bam")
                    command1 = "samtools view -bS "+samfile_in+ " > " + samfile_out
                    showStatus(command1)
                    os.system(command1)
                    os.remove(samfile_in) #remove temp file
                    
                    # sort reads
                    samfile_in=samfile_out
                    samfile_out=samfile_in.replace("new", "sorted").replace(".bam", "") #the sort will create '.bam' for extension
                    command2 = "samtools sort " + samfile_in + "  " + samfile_out
                    showStatus(command2)
                    os.system(command2)
                    os.remove(samfile_in) #remove temp file
                    
                    # remove duplicates
                    samfile_in=samfile_out+".bam"
                    samfile_out=samfile_in.replace("sorted", "rmdup")
                    command3 = "samtools rmdup -S " + samfile_in + " " + samfile_out
                    showStatus(command3)
                    os.system(command3)
                    os.remove(samfile_in) #remove temp file

                    # convert back to sam format
                    samfile_in=samfile_out
                    samfile_out=samfile_in.replace(".bam", ".sam")
                    command4 = "samtools view " + samfile_in + " > " + samfile_out
                    showStatus(command4)
                    os.system(command4)
                    #os.remove(samfile_in) #remove temp file

                    if folders[r].count('.') > 1 or ('__' not in folders[r]) :
                            dotloc = folders[r].index('.')
                    else:
                            dotloc = folders[r].index('__')
                    samplename = folders[r][:dotloc]
                    mapped = 0

                    samfile = samfile_out
                    f_in = open(samfile, "r")
                    process_index = 0
                    process_start_time = time.time()
                    line = f_in.readline().replace("\n", "")

                    while (line):
                            if (process_index % 200) == 0:
                                time_used_sec = str(round(time.time() - process_start_time,2))
                                info = " reads processed: " + str(process_index) + ", time used: " + time_used_sec +" sec"
                                showStatus(info)
                            items=line.split("\t")
                            r_qname = items[0]
                            r_flag = int(items[1])
                            r_rname = items[2]
                            r_pos = int(items[3])
                            r_mapq = int(items[4])
                            r_cigar = items[5]
                            r_rnext = items[6]
                            r_pnext = int(items[7])
                            r_tlen = int(items[8])
                            r_seq = items[9]
                            r_qual = items[10]
                            digits = re.findall('\d+', r_cigar)
                            letters = re.findall('\D', r_cigar)
                            r_tags=""
                            for tag in range(11, len(items)):
                                r_tags = r_tags + items[tag]+" "
                            line = f_in.readline().replace("\n", "")
                            process_index += 1 ## debug
                            r_mapped_len=0
                            skip_reads=False
                            print_reads=False
                            s_cnt = 0
                            for i in letters:
                                if i in "X=":
                                    skip_reads = True
                                    break
                            if skip_reads == True: #skip reads with '=' or 'x'
                                continue
                            #print("process_index = ", process_index, " mapped = ", mapped) ## debug
                            #logInfo("process_index = "+ str(process_index)+ " mapped = "+ str(mapped)) ## debug
                            if r_pos > 0: #mapped reads
                                    #print ('query name = ' + r_qname)
                                    #print ('reference name = ' + r_rname)
                                    ind = str(r_rname).index('-')
                                    genotype = str(r_rname)[ind+1:]
                                    #print ('cigar = ', r_cigar, 'mapping length = ', r_mapped_len, " letters len = ", len(letters))
                                    #print ('read sequence  = ' + r_seq) #read sequence bases, included soft clipped bases
                                    #print ('aligned sequence = ' + read.query_alignment_sequence) #aligned portion of read, excludes flanking bases that were soft clipped
                                    cigar_len = len(letters)
                                    r_mapped_len=0
                                    r_seq_new = ""
                                    r_start=0
                                    r_end=0
                                    for i in range(cigar_len):
                                        if letters[i]=='M':
                                            r_end=r_start+int(digits[i])
                                            r_seq_new = r_seq_new+r_seq[r_start:r_end] #copy match
                                            r_start=r_end
                                            r_mapped_len +=  int(digits[i])
                                        if letters[i]=='I':
                                            r_end=r_start+int(digits[i]) #skip inserted
                                            r_start=r_end
                                        if letters[i]=='S':
                                            r_end=r_start+int(digits[i]) #skip 'skip'
                                            r_start=r_end
                                        if letters[i]=='D':
                                            for ii in range(int(digits[i])):
                                                r_seq_new = r_seq_new+' '
                                            r_mapped_len +=  int(digits[i])
                                    try:
                                        seq_record = HBV_data_dict[r_rname]
                                        reference = str(seq_record.seq)

                                        # FSS ----------------------------------
                                        if first_reference == None:
                                            # get first reference name
                                            first_reference = r_rname
                                            add_read_locations = True
                                            add_mut_count = 1
                                            read_ref_length = len(reference)
                                        else:
                                            if first_reference == r_rname:
                                                #only get count the reference is same as the first reference
                                                add_read_locations = True
                                                add_mut_count=1
                                            else:
                                                add_read_locations = False
                                                add_mut_count = 0

                                    except:
                                        print("record id not found = ", r_rname)
                                        logInfo ('\n!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!\n')
                                        logInfo ('HBV reference not found.')
                                        logInfo ('You either: ')
                                        logInfo (' (1) copy HBV reference to DriverFinder folder and ')
                                        logInfo ('      name it "HBV_complete.fa", or')
                                        logInfo (' (2) use "Select HBV Reference File" to get HBV reference file.')
                                        logInfo ('\n!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!\\nn')
                                        if (parent != None):
                                            parent.enableOptions()
                                        return
                                        continue
                                    r_alignment = reference[r_pos - 1:(r_pos + r_mapped_len - 1)]
                                    referenceseq = reference[r_pos - 1:(r_pos + r_mapped_len - 1)]
                                    alignedseq = r_seq_new

                                    # FSS ---------------------------------------
                                    if (add_read_locations == True):
                                        for read_ndx in range(r_pos, r_pos + r_mapped_len):
                                            read_locations[read_ndx]+=1
                                    """
                                    r_seq:        read sequence
                                    r_seq_new:    read sequence with 'skip' and 'inserted' remove and ' ' added for 'deleted'
                                    alignedseq:   read sequence with 'skip' and 'inserted' remove and ' ' added for 'deleted'
                                    reference:    entire reference sequence from the reference file
                                    referenceseq: part of reference which has same length as alignedseq
                                    r_alignment:  part of reference which has same length as alignedseq
                                    """
                                    #find mutation
                                    m_loc=[]
                                    m_val=[]
                                    for i in range(len(r_alignment)):
                                        alignment_temp = r_alignment[i].upper()
                                        viral_temp = r_seq_new[i]
                                        if (alignment_temp != viral_temp) and (viral_temp != ' '):
                                            m_loc.append(i+r_pos)
                                            val = alignment_temp+'/'+viral_temp
                                            m_val.append(val)

                                    """
                                    # check for non regular data
                                    r_temp=r_alignment.replace("a","").replace("c","").replace("g","").replace("t","")
                                    if (len(r_temp) > 2):
                                        print("FSS unknown ----",r_temp,"-----", m_val)
                                        print_reads = True
                                    """            
                                    if print_reads == True:
                                        print("FSS CIGAR = ", r_cigar)
                                        print('read sequence  = ' + r_seq) #read sequence bases, included soft clipped bases
                                        print('new  sequence  = ' + r_seq_new) #read sequence bases, included soft clipped bases
                                        print('read alignment = ' + r_alignment)

                                    """
                                    reference: entire reference sequence from the reference file
                                    alignedseq: sequence from the read
                                    referenceseq:???
                                    """

                                    #referenceseq = str(read.get_reference_sequence())
                                    #print ('query alignment coordinates = ' + str(read.query_alignment_start) + ', ' + str(read.query_alignment_end))
                                    #print ('reference coordinates = ' + str(read.reference_start) + ', ' + str(read.reference_end))

                                    # check if known mutations found on the current read
                                    readcords = (r_pos, r_pos+r_mapped_len)
                                    for x in range(0, len(knownmutations)):
                                            string = knownmutations[x]
                                            mut_num = string.count('/') + 1
                                            completed = []
                                            digits = re.findall('\d+', string)
                                            letters = re.findall('\D', string)
                                            letters = [x for x in letters if x != '/']
                                            check = True

                                            #if known mutation location is not within current read, no check is needed
                                            for l in digits:
                                                    if int(l) < readcords[0] or int(l) > readcords[1]:
                                                            check = False
                                            if check == True:
                                                    #print ('Checking for mutation at ' + ', '.join(digits) + ' ...')
                                                    place = 0
                                                    for i in range(0, mut_num):
                                                            nt1 = str(reference[int(digits[i]) - 1]).upper()
                                                            try:
                                                                    nt2 = str(alignedseq[int(digits[i]) - readcords[0] - 1]).upper()
                                                            except:
                                                                    print ('?? position not found in aligned seq ??')
                                                                    print (alignedseq)
                                                                    print (int(digits[i]) - readcords[0] - 1)
                                                                    break
                                                            mut_nt1 = letters[place]
                                                            mut_nt2 = letters[place+1]
                                                            #print('nt1 = '+nt1+', nt2 = '+nt2+', mut_nt1 = '+mut_nt1+', mut_nt2 = '+mut_nt2)
                                                            if nt1 != nt2 and nt1 in IUPAC_nt[mut_nt1] and nt2 in IUPAC_nt[mut_nt2]:
                                                                    debug_count += 1
                                                                    if (debug_count < 0): #debug
                                                                        print('mutation nt1 = '+nt1+', nt2 = '+nt2+', mut_nt1 = '+mut_nt1+', mut_nt2 = '+mut_nt2)
                                                                        if (nt1 == mut_nt1) and (nt2 == mut_nt2):
                                                                            print('mutation found! ---- same mutation ----------')
                                                                    completed.append(mut_num)
                                                                    """
                                                                    if i == mut_num - 1:
                                                                            print (str(len(completed))+' mutations found out of '+ str(mut_num))
                                                                    """
                                                                    refstring = ''
                                                                    if reflinks[x].count(',') > 0:
                                                                            refstring = [x.strip() for x in reflinks[x].split(',')]
                                                                            for ref in range(0, len(refstring)):
                                                                                    if 'ncbi' in refstring[ref]:
                                                                                            refstring[ref] = 'PMC' + str(re.findall('\d+', refstring[ref]))[2:-2]
                                                                            refstring = str(refstring)[2:-2]
                                                                            refstring = refstring.replace("'", "")
                                                                    else:
                                                                            refstring = reflinks[x].strip()
                                                                            refstring = refstring.replace("'", "")
                                                                    if len(completed) == 0:
                                                                            results.append(OrderedDict([('Read Name', r_qname), ('Genotype', genotype), \
                                                                                                        ('Mutation', knownmutations[x]), \
                                                                                                        ('Functional Significance', function[x]), \
                                                                                                        ('Reference(s)', refstring),\
                                                                                                        ('Read Sequence', r_seq), \
                                                                                                        ('Reference Sequence', referenceseq), \
                                                                                                        ('Reference Coordinates', (r_pos, r_pos+r_mapped_len)), \
                                                                                                        ('Query Alignment Coordinates', (r_pos, r_pos+r_mapped_len))]))
                                                                    else:
                                                                            temp = knownmutations[x].split('/')
                                                                            for c in completed:
                                                                                    results.append(OrderedDict([('Sample', samplename), ('Read Name', r_qname), ('Genotype', genotype), \
                                                                                                                ('Mutation', temp[c-1]), ('Functional Significance', function[x]), \
                                                                                                                ('Reference(s)', refstring),\
                                                                                                                ('Read Sequence', r_seq), \
                                                                                                                ('Reference Sequence', referenceseq), \
                                                                                                                ('Reference Coordinates', (r_pos, r_pos+r_mapped_len)), \
                                                                                                                ('Query Alignment Coordinates', (r_pos, r_pos+r_mapped_len))]))			
                                                            """
                                                            elif i == mut_num - 1:
                                                                    print (str(len(completed))+' mutations found out of '+ str(mut_num)+'\n')
                                                            """
                                                            place += 2
                                    
                                    count = 0
                                    i = 0

                                    #print ('Other mutations found:')
                                    for l in referenceseq:
                                            if (len(referenceseq) != len(alignedseq)):
                                                    print("ref = ", referenceseq)
                                                    print("alg = ", alignedseq)
                                                    sys.stdout.flush()
                                                    break
                                            if (alignedseq[i] == 'N'):
                                                    i=i+1
                                                    continue
                                            if (alignedseq[i] == ' '):
                                                    # FSS --------------------------------
                                                    # decrease count on location 'i+readcords[0]'
                                                    read_locations[i+readcords[0]] -= add_mut_count
                                                    i=i+1
                                                    continue
                                            if (alignedseq[i] == l.upper()):
                                                    i=i+1
                                                    continue
                                            if (l.islower() and l != 'n') or l == 'c' or l == 'g' or l == 'a' or l == 't':
                                                    try:
                                                            mut = l.upper() + str(i+readcords[0]) + alignedseq[i]
                                                            #print (mut)
                                                    except:
                                                            mut = l.upper() + str(i+readcords[0]) + '?'
                                                            #print (mut)
                                                    # foundmutations.append(OrderedDict([('Sample', samplename), ('Read Name', read.query_name), ('Genotype', genotype), ('Mutation', mut), ('Read Sequence', read.query_sequence), ('Reference Sequence', referenceseq), ('Reference Coordinates', (read.reference_start, read.reference_end)), ('Query Alignment Coordinates', (read.query_alignment_start, read.query_alignment_end))]))
                                                    
                                                    if in_dictlist(i+readcords[0]) > -1:
                                                            index = in_dictlist(i+readcords[0])
                                                            orderedmutations[index]['Count'] += 1

                                                            # FSS --------------------------------
                                                            # increase mutation count on location 'i+readcords[0]' 
                                                            read_locations_counts[i+readcords[0]] += add_mut_count
                                                            read__types[i+readcords[0]]=l.upper()
                                                            
                                                            try:
                                                                    # before: a= t= c= g= ?= 
                                                                    temp = orderedmutations[index]['WildType']
                                                                    #print (temp)
                                                                    temp[l.upper()] += 1
                                                                    #print (temp)
                                                                    orderedmutations[index]['WildType'] = temp
                                                            except:
                                                                    temp = orderedmutations[index]['WildType']
                                                                    temp['?'] += 1
                                                                    orderedmutations[index]['WildType'] = temp
                                                            temp1 = orderedmutations[index]['Mutant']
                                                            temp1[alignedseq[i]] += 1
                                                            orderedmutations[index]['Mutant']= temp1
                                                            #add genotype
                                                            temp = {genotype: 1}
                                                            orderedmutations[index]['Genotypes'] = { k: temp.get(k, 0) + orderedmutations[index]['Genotypes'].get(k, 0) for k in set(temp) | set(orderedmutations[index]['Genotypes']) }
                                                            #add sample
                                                            samples_list = orderedmutations[index]['Samples']
                                                            if samplename not in samples_list:
                                                                    orderedmutations[index]['Samples'].append(samplename)
                                                            #add read name
                                                            # orderedmutations[index]['Reads'].append(read.query_name)
                                                            count += 1
                                                    else:
                                                            try: 
                                                                    temp = {'A':0, 'T':0, 'C':0, 'G':0, '?':0}
                                                                    temp[l.upper()] = 1
                                                                    temp1 = {'A':0, 'T':0, 'C':0, 'G':0, '?':0}
                                                                    temp1[alignedseq[i]] = 1
                                                                    orderedmutations.append(OrderedDict([('Count', 1), ('Location', i+readcords[0]), ('WildType', temp), ('Mutant', temp1), ('Genotypes', {genotype: 1}), ('Samples', [samplename])]))
                                                            except:
                                                                    temp = {'A':0, 'T':0, 'C':0, 'G':0, '?':1}
                                                                    temp1 = {'A':0, 'T':0, 'C':0, 'G':0, '?':0}
                                                                    temp1[alignedseq[i]] = 1
                                                                    orderedmutations.append(OrderedDict([('Count', 1), ('Location', i+readcords[0]), ('WildType', temp), ('Mutant', temp1), ('Genotypes', {genotype: 1}), ('Samples', [samplename])]))

                                                            # FSS --------------------------------
                                                            # increase mutation count on location 'i+readcords[0]' 
                                                            read_locations_counts[i+readcords[0]] += add_mut_count
                                                            read__types[i+readcords[0]]=l.upper()

                                            i += 1
                                    #print ('\n==========================================================\n')
                                    
                                    mapped += 1


                    #print (str(len(foundmutations)) + ' mutations found total')
                    f_in.close()
                    print("count = ", count)
                    print (str(mapped) + ' mapped reads total\n')
                    logInfo(str(mapped) + ' mapped reads total\n')
                    logInfo("Updating mutation database.....")

                    t1_d = time.time()

                    if (mapped > 0):
                        if not os.path.isfile('MutationDatabase.xlsx'):
                                print ('Creating Mutation Database file...')
                                writer = ExcelWriter('MutationDatabase.xlsx')
                                headers = ['Sample', 'Read Name', 'Genotype', 'Mutation', 'Functional Significance', 'Reference(s)', 'Read Sequence', 'Reference Sequence', 'Reference Coordinates', 'Query Alignment Coordinates']
                                results_df = pd.DataFrame(results, columns = headers)
                                results_df.to_excel(writer, 'Known Mutations', index = False)
                                # headers = ['Sample', 'Read Name', 'Genotype', 'Mutation', 'Read Sequence', 'Reference Sequence', 'Reference Coordinates', 'Query Alignment Coordinates']
                                # foundmuts_df = pd.DataFrame(foundmutations, columns = headers)
                                # foundmuts_df.to_excel(writer, 'Found Mutations', index = False)
                                orderedmutations = sorted(orderedmutations, key=lambda k: k['Count'], reverse = True) 
                                headers = ['Count', 'Location', 'WildType', 'Mutant', 'Genotypes', 'Samples']
                                orderedmuts_df = pd.DataFrame(orderedmutations, columns = headers)
                                orderedmuts_df.to_excel(writer, 'Found Mutations', index = False)
                                writer.save()
                                for x_ndx in range(len(orderedmutations)):
                                    x_loc = orderedmutations[x_ndx]['Location']
                                    x_mutant=str(orderedmutations[x_ndx]['Mutant'])
                                    mutant_max_type, mutant_max_count = find_mutant_max(read__types[x_loc], x_mutant)
                                    read_max_type[x_loc]=mutant_max_type
                                    read_max_count[x_loc]=mutant_max_count
                                    #print(x_loc, x_mutant," = ", read__types[x_loc], x_loc, mutant_max_type, mutant_max_count)
                        else:
                                print ('Appending results to Mutation Database... \n')
                                book = load_workbook('MutationDatabase.xlsx')
                                writer = ExcelWriter('MutationDatabase.xlsx', engine = 'openpyxl')
                                writer.book = book
                                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                                # Convert 'Known Mutations' sheet to dataframe
                                df = pd.read_excel('MutationDatabase.xlsx', header = 0)
                                headers = ['Sample', 'Read Name', 'Genotype', 'Mutation', 'Functional Significance', 'Reference(s)', 'Read Sequence', 'Reference Sequence', 'Reference Coordinates', 'Query Alignment Coordinates']
                                known_muts_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
                                for letter in range(0, len(headers)):
                                        templist = df[headers[letter]].tolist()
                                        known_muts_cols[letter] = templist

                                known_muts_cols = pd.DataFrame({'Sample': known_muts_cols[0], 'Read Name': known_muts_cols[1], 'Genotype': known_muts_cols[2], 'Mutation': known_muts_cols[3], 'Functional Significance': known_muts_cols[4], 'Reference(s)': known_muts_cols[5], 'Read Sequence': known_muts_cols[6], 'Reference Sequence': known_muts_cols[7], 'Reference Coordinates': known_muts_cols[8], 'Query Alignment Coordinates': known_muts_cols[9]}, columns = headers)

                                # Convert 'Found Mutations' sheet to dataframe
                                # df = pd.read_excel('MutationDatabase.xlsx', header = 0, sheetname = 1)
                                # headers = ['Sample', 'Read Name', 'Genotype', 'Mutation', 'Read Sequence', 'Reference Sequence', 'Reference Coordinates', 'Query Alignment Coordinates']
                                # found_muts_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
                                # for letter in range(0, len(headers)):
                                # 	templist = df[headers[letter]].tolist()
                                # 	found_muts_cols[letter] = templist

                                # found_muts_cols = pd.DataFrame({'Sample': found_muts_cols[0], 'Read Name': found_muts_cols[1], 'Genotype': found_muts_cols[2], 'Mutation': found_muts_cols[3], 'Read Sequence': found_muts_cols[4], 'Reference Sequence': found_muts_cols[5], 'Reference Coordinates': found_muts_cols[6], 'Query Alignment Coordinates': found_muts_cols[7]}, columns = headers)

                                # Convert'Found Mutations (1)' sheet to dataframe
                                df = pd.read_excel('MutationDatabase.xlsx', header = 0, sheetname = 1)
                                headers = ['Count', 'Location', 'WildType', 'Mutant', 'Genotypes', 'Samples']
                                found_muts_1_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
                                for letter in range(0, len(headers)):
                                        templist = df[headers[letter]].tolist()
                                        #print (templist)
                                        found_muts_1_cols[letter] = templist
                                # found_muts_1_cols = pd.DataFrame({'Count': found_muts_1_cols[0], 'Location': found_muts_1_cols[1], 'WildType': found_muts_1_cols[2], 'Mutant': found_muts_1_cols[3], 'Genotypes': found_muts_1_cols[4], 'Sample': found_muts_1_cols[5], 'Reads': found_muts_1_cols[6]}, columns = headers)

                                for loc in range(0, len(found_muts_1_cols[1])):
                                        #print (int(found_muts_1_cols[1][loc]))
                                        if in_dictlist(int(found_muts_1_cols[1][loc])) > -1:
                                                #print ('in dict list!')

                                                index = in_dictlist(int(found_muts_1_cols[1][loc]))
                                                #add count
                                                orderedmutations[index]['Count'] += found_muts_1_cols[0][loc]
                                                #add before
                                                letters = literal_eval(str(found_muts_1_cols[2][loc]))
                                                orderedmutations[index]['WildType'] = { k: letters.get(k, 0) + orderedmutations[index]['WildType'].get(k, 0) for k in set(letters) | set(orderedmutations[index]['WildType']) }
                                                #add after
                                                letters = literal_eval(str(found_muts_1_cols[3][loc]))
                                                orderedmutations[index]['Mutant'] = { k: letters.get(k, 0) + orderedmutations[index]['Mutant'].get(k, 0) for k in set(letters) | set(orderedmutations[index]['Mutant']) }
                                                #add genotypes
                                                letters = literal_eval(str(found_muts_1_cols[4][loc]))
                                                orderedmutations[index]['Genotypes'] =  { k: letters.get(k, 0) + orderedmutations[index]['Genotypes'].get(k, 0) for k in set(letters) | set(orderedmutations[index]['Genotypes']) }
                                                #add sample name
                                                for x in convertstring(found_muts_1_cols[5][loc]):
                                                        if x not in orderedmutations[index]['Samples']:
                                                                orderedmutations[index]['Samples'].append(x)
                                                #add read name
                                                # for x in convertstring(found_muts_1_cols[6][loc]):
                                                # 	if x not in orderedmutations[index]['Reads']:
                                                # 		orderedmutations[index]['Reads'].append(x)
                                                # orderedmutations[index]['Reads'] += convertstring(found_muts_1_cols[6][loc])
                                                # orderedmutations[index]['Reads'] = ', '.join(map(str, orderedmutations[index]['Reads']))
                                        else:
                                                letters = literal_eval(str(found_muts_1_cols[2][loc]))
                                                temp_wildtype = { k: letters.get(k, 0) for k in set(letters)}
                                                letters = literal_eval(str(found_muts_1_cols[3][loc]))
                                                temp_mutant = { k: letters.get(k, 0) for k in set(letters)}
                                                letters = literal_eval(str(found_muts_1_cols[4][loc]))
                                                temp_genotype = { k: letters.get(k, 0) for k in set(letters)}
                                                temp_samples=[]
                                                for x in convertstring(found_muts_1_cols[5][loc]):
                                                        if x not in temp_samples:
                                                                temp_samples.append(x)
                                                orderedmutations.append({'Count': found_muts_1_cols[0][loc], 'Location': found_muts_1_cols[1][loc],
                                                                         'WildType': temp_wildtype, 'Mutant': temp_mutant,
                                                                         'Genotypes': temp_genotype, 'Samples': temp_samples})
                                                """
                                                orderedmutations.append({'Count': found_muts_1_cols[0][loc], 'Location': found_muts_1_cols[1][loc],
                                                                         'WildType': orderedmutations[index]['WildType'], 'Mutant': orderedmutations[index]['Mutant'],
                                                                         'Genotypes': orderedmutations[index]['Genotypes'], 'Samples': orderedmutations[index]['Samples']})
                                                """

                                # headers = ['Count', 'Location', 'WildType', 'Mutant', 'Genotypes', 'Sample', 'Reads']
                                # found_muts_1_cols = pd.DataFrame({'Count': found_muts_1_cols[0], 'Location': found_muts_1_cols[1], 'WildType': found_muts_1_cols[2], 'Mutant': found_muts_1_cols[3], 'Genotypes': found_muts_1_cols[4], 'Sample': found_muts_1_cols[5], 'Reads': found_muts_1_cols[6]}, columns = headers)

                                results_df = pd.DataFrame(results)
                                # foundmuts_df = pd.DataFrame(foundmutations)
                                orderedmuts_df = pd.DataFrame(orderedmutations)

                                frame1 = pd.concat([known_muts_cols, results_df])
                                frame1 = frame1.drop_duplicates(['Genotype'])
                                frame1 = frame1.sort_values(by = 'Mutation')
                                # frame2 = pd.concat([found_muts_cols, foundmuts_df])
                                # frame2 = frame2.sort_values(by = 'Mutation')
                                orderedmutations = sorted(orderedmutations, key = lambda k: k['Count'], reverse = True)
                                for x_ndx in range(len(orderedmutations)):
                                    x_loc = orderedmutations[x_ndx]['Location']
                                    x_mutant=str(orderedmutations[x_ndx]['Mutant'])
                                    mutant_max_type, mutant_max_count = find_mutant_max(read__types[x_loc], x_mutant, x_loc == 2612)
                                    read_max_type[x_loc]=mutant_max_type
                                    read_max_count[x_loc]=mutant_max_count
                                    #print(x_loc, x_mutant," = ", read__types[x_loc], x_loc, mutant_max_type, mutant_max_count)
                                orderedmutations = pd.DataFrame(orderedmutations)
                                orderedmutations.sort_values('Count', ascending=False)
                                frame1.to_excel(writer, 'Known Mutations', index = False)
                                # frame2.to_excel(writer, 'Found Mutations', index = False)
                                orderedmutations.to_excel(writer, 'Found Mutations', index = False)

                                writer.save()
                                print ('Results Saved to MutationDatabase.xlsx! \n')

                    # Reset!
                    foundmutations = []
                    orderedmutations = []
                    results = []
                    t2=time.time()
                    print("file ", filename, " processed")
                    print("   time used = ", round(t2-t1, 2), " (find mut time = ", round(t1_d-t1,2), "  database time = ", round(t2-t1_d,2),")")
                    logInfo("file "+ filename +" processed")
                    logInfo("  time used = "+ str(round(t2-t1,2))+ " (find mut time = "+ str(round(t1_d-t1,2))+ "  database time = "+ str(round(t2-t1_d,2))+")")
                    t_m = t_m + (t1_d - t1)
                    t_d = t_d + (t2 - t1_d)
                    t1=time.time()
                    sys.stdout.flush()
                    #os.remove(samfile_out)

    # FSS ---------------------------------------
    print(" save frequency in a CSV file 'Mutation_Frequency.csv'")
    c_t = datetime.datetime.now()
    c_t_str=str(c_t.year)+"_"+str(c_t.month)+"_"+str(c_t.day)+"_"
    c_t_str=c_t_str + str(c_t.hour)+"_"+str(c_t.minute)+"_"+str(c_t.second)
    samplename_path, samplename_name=os.path.split(samplename)
    f = open("Mutation_Frequency_" +samplename_name+"_"+c_t_str + ".csv", "w")
    f.write("location,read count,mutation count,pct,type,max count,max type,mutation,mut pct\n")
    for ndx in range(1, read_ref_length+1):
        f.write(str(ndx)+","+str(read_locations[ndx])+","+str(read_locations_counts[ndx])+",")
        if read_locations[ndx] == 0:
            # if no read in this location, use '-1' for percentage
            f.write("-1"+",")
        else:
            # calculate percentage 
            f.write(str(100*read_locations_counts[ndx]/read_locations[ndx])+",")
        f.write(read__types[ndx]+","+str(read_max_count[ndx])+","+read_max_type[ndx]+",")
        if read_locations[ndx] == 0:
            # if no read in this location, use '-1' for percentage
            f.write(" ,"+"-1")
        else:
            if int(read_max_count[ndx])>0:
                mutation_str=str(read__types[ndx])+str(ndx)+str(read_max_type[ndx])
                mutation_str=mutation_str.replace(" ","")
            else:
                mutation_str=" "
            f.write(mutation_str+",")
            # calculate percentage 
            f.write(str(100*int(read_max_count[ndx])/read_locations[ndx]))
        f.write("\n")
    f.close()


    shutil.move('MutationDatabase.xlsx',
                'MutationDatabase_'+samplename_name+"_"+c_t_str +'.xlsx')
    
    print('Time Elapsed: ' + str(round(timer() - start,2)), "(mut time = ",round(t_m,2), "  database time = ",round(t_d,2),")")
    logInfo('\nComplete, Time Elapsed: ' + str(round(timer() - start,2))+" ( mut = "+str(round(t_m,2))+" database = " +str(round(t_d,2))+" )")
    showStatus(" ") # clear status line

    sys.stdout.flush()

    sys.stdout = old_stdout
    if log_file != None:
        log_file.close()

    if (parent != None):
        parent.enableOptions()

